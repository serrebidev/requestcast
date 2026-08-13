from __future__ import annotations

import base64
import hashlib
import hmac
import itertools
import json
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import unicodedata
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse
from xml.etree import ElementTree

import requests
from flask import Flask, abort, flash, redirect, render_template, request, session, url_for
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from mutagen import File as MutagenFile
from mutagen.aiff import AIFF
from mutagen.flac import FLAC, Picture
from mutagen.id3 import APIC, COMM, ID3, TALB, TDRC, TIT2, TPE1, TPOS, TRCK, TXXX
from mutagen.mp4 import MP4, MP4Cover
from mutagen.wave import WAVE
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.middleware.proxy_fix import ProxyFix
from ytmusicapi import YTMusic

from . import config, deezer, musicdl_source, permissions, tools


HTTP_TIMEOUT = (10, 30)
MAX_COLLECTION_TRACKS = 100
# Whole-library lists are normal, so the ceilings are set by what parsing actually costs,
# not by a cautious round number. On a 2020-era machine 250,000 rows takes about a second
# from TXT, eleven seconds from XLSX, and a PDF page costs roughly three milliseconds.
MAX_IMPORT_ENTRIES = 250_000
MAX_IMPORT_TRACKS = 250_000
MAX_UPLOAD_BYTES = 64 * 1024 * 1024
# Raw rows tolerated before deduplication, so a file far over the limit fails fast
# instead of being read into memory in full first.
MAX_RAW_IMPORT_ENTRIES = MAX_IMPORT_ENTRIES * 2
MAX_PDF_PAGES = 5_000
# Artist-only import lines take the artist's whole catalogue unless the uploader picks a cap.
IMPORT_ARTIST_TRACK_CHOICES = {"all": 0, "100": 100, "50": 50, "25": 25, "10": 10}
MAX_ARTIST_ALBUMS = 300
# How many releases or videos one browse page offers per section.
MAX_BROWSE_ITEMS = 200
BROWSABLE_KINDS = {"artist", "channel"}
# What the search page offers as "results per source". The saved preference is the
# starting point; each search can ask for a different number from this list.
SEARCH_LIMIT_CHOICES = (10, 25, 50, 100, 200)
# Deezer returns at most 100 rows per call, so larger limits are paged.
DEEZER_PAGE_SIZE = 100
# yt-dlp asks YouTube as one of several clients. When a download comes back 403 or
# "unavailable" during a bulk run, the next attempt asks as a different client, which
# is what usually clears it. The first attempt leaves yt-dlp to its own default.
YTDLP_PLAYER_CLIENTS = ("", "tv,web_safari", "ios,mweb", "android,web")
# Failures that mean "the site is pushing back", not "this track does not exist".
# A run of these during a channel, discography, or playlist import is rate limiting:
# the same tracks download fine once the queue slows down.
RATE_LIMIT_SIGNS = (
    "403", "429", "too many requests", "rate limit", "rate-limit", "throttl",
    "sign in to confirm", "not a bot", "video unavailable", "unable to download",
    "temporarily unavailable", "please try again later", "connection reset",
    "read timed out", "timed out", "failed to extract", "unavailable videos",
)
# Consecutive rate-limited failures before the job pauses to let the site settle.
RATE_LIMIT_STREAK = 2
# Playlists people already have. A playlist names files this machine may not hold, so
# each entry is treated as a request for that artist and title, not as a local file.
PLAYLIST_EXTENSIONS = {
    ".m3u", ".m3u8", ".pls", ".xspf", ".wpl", ".asx", ".cue", ".fpl", ".fb2k-playlist",
}
IMPORT_EXTENSIONS = {".txt", ".xlsx", ".pdf"} | PLAYLIST_EXTENSIONS
AUDIO_FILE_EXTENSIONS = {
    ".aac", ".aif", ".aiff", ".ape", ".flac", ".m4a", ".mp2", ".mp3", ".mp4", ".oga",
    ".ogg", ".opus", ".wav", ".wma", ".wv",
}
AZURACAST_AUDIO_EXTENSIONS = {
    ".aac", ".aif", ".aiff", ".flac", ".m4a", ".mp2", ".mp3", ".mp4",
    ".oga", ".ogg", ".opus", ".wav", ".wma",
}

# Set by the test suite and by one-off tooling. The download worker, the automatic tool
# installer, and the update checker all stay out of the way when it is set, so nothing
# reaches the network behind a test's back.
WORKER_DISABLED = (
    os.environ.get("REQUESTCAST_DISABLE_WORKER") == "1"
    or os.environ.get("ADDTO_DISABLE_WORKER") == "1"
)

app = Flask(__name__)
app.permanent_session_lifetime = timedelta(hours=12)
app.config.update(
    SESSION_COOKIE_NAME="requestcast_session",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    MAX_CONTENT_LENGTH=MAX_UPLOAD_BYTES,
)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

SETTINGS: dict[str, Any] = {}
STATE_DIR = Path(".")
DOWNLOAD_DIR = Path(".")
MEDIA_DIR = Path(".")
DB_PATH = Path("jobs.sqlite3")
YTDLP = ""
FFMPEG = ""
FFPROBE = ""
DENO = ""
AZURACAST_ENABLED = False
AZURACAST_API_BASE = ""
AZURACAST_API_KEY = ""
STATION_ID = config.DEFAULT_STATION_ID
REQUEST_PLAYLIST_ID = ""
UPLOAD_DIR = config.DEFAULT_UPLOAD_DIRECTORY
SECRET_KEY = ""
PASSWORD_SALT = b""
PASSWORD_HASH = b""
ADMIN_PASSWORD_SALT = b""
ADMIN_PASSWORD_HASH = b""
DEEZER: deezer.DeezerClient | None = None
DEEZER_ERROR = ""
MUSICDL_ENABLED = False
MUSICDL_SOURCES: list[str] = []
SEARCH_RESULT_LIMIT = 50
SEARCH_MUSICDL = False
DOWNLOAD_RETRIES = 2
DOWNLOAD_RETRY_DELAY = 20
DOWNLOAD_GAP_SECONDS = 2
RATE_LIMIT_COOLDOWN = 180
JOB_RETRY_LIMIT = 1
AUTO_UPDATE_TOOLS = True
AUTO_UPDATE_INTERVAL_HOURS = 24
signer: URLSafeTimedSerializer | None = None


def station_api(path: str) -> str:
    return f"{AZURACAST_API_BASE}/station/{STATION_ID}{path}"


def apply_settings(new_settings: dict[str, Any] | None = None) -> None:
    """Load settings into module state. Safe to call again after the setup page saves."""
    global SETTINGS, STATE_DIR, DOWNLOAD_DIR, MEDIA_DIR, DB_PATH, YTDLP, FFMPEG, FFPROBE, DENO
    global AZURACAST_ENABLED, AZURACAST_API_BASE, AZURACAST_API_KEY, STATION_ID
    global REQUEST_PLAYLIST_ID, UPLOAD_DIR, SECRET_KEY, PASSWORD_SALT, PASSWORD_HASH
    global ADMIN_PASSWORD_SALT, ADMIN_PASSWORD_HASH, signer
    global DEEZER, DEEZER_ERROR
    global MUSICDL_ENABLED, MUSICDL_SOURCES
    global SEARCH_RESULT_LIMIT, SEARCH_MUSICDL
    global DOWNLOAD_RETRIES, DOWNLOAD_RETRY_DELAY, DOWNLOAD_GAP_SECONDS, RATE_LIMIT_COOLDOWN
    global JOB_RETRY_LIMIT, AUTO_UPDATE_TOOLS, AUTO_UPDATE_INTERVAL_HOURS

    SETTINGS = new_settings if new_settings is not None else config.load()
    STATE_DIR = Path(SETTINGS["state_dir"])
    DOWNLOAD_DIR = Path(SETTINGS["download_dir"])
    DB_PATH = STATE_DIR / "jobs.sqlite3"
    YTDLP = tools.find_tool("yt-dlp", SETTINGS.get("ytdlp_path", ""))
    FFMPEG = tools.find_tool("ffmpeg", SETTINGS.get("ffmpeg_path", "")) or "ffmpeg"
    FFPROBE = tools.find_tool("ffprobe", SETTINGS.get("ffprobe_path", "")) or "ffprobe"
    # yt-dlp hands YouTube's JavaScript challenges to Deno. yt-dlp finds Deno on PATH by
    # itself; this is for our own copy in the tools folder, which is not on PATH.
    DENO = tools.find_tool("deno", SETTINGS.get("deno_path", ""))

    AZURACAST_ENABLED = bool(SETTINGS.get("azuracast_enabled"))
    AZURACAST_API_BASE = str(SETTINGS.get("azuracast_api_base", "")).rstrip("/")
    AZURACAST_API_KEY = str(SETTINGS.get("azuracast_api_key", ""))
    STATION_ID = str(SETTINGS.get("azuracast_station_id") or config.DEFAULT_STATION_ID)
    REQUEST_PLAYLIST_ID = str(SETTINGS.get("azuracast_request_playlist_id", ""))
    UPLOAD_DIR = str(SETTINGS.get("azuracast_upload_dir") or config.DEFAULT_UPLOAD_DIRECTORY)
    # Without AzuraCast the downloads folder is the final destination for finished audio.
    MEDIA_DIR = Path(SETTINGS["azuracast_media_dir"]) if (
        AZURACAST_ENABLED and SETTINGS.get("azuracast_media_dir")
    ) else DOWNLOAD_DIR

    SECRET_KEY = str(SETTINGS.get("secret_key", ""))
    PASSWORD_SALT = bytes.fromhex(SETTINGS["password_salt"]) if SETTINGS.get("password_salt") else b""
    PASSWORD_HASH = bytes.fromhex(SETTINGS["password_hash"]) if SETTINGS.get("password_hash") else b""
    ADMIN_PASSWORD_SALT = bytes.fromhex(SETTINGS["admin_password_salt"]) if SETTINGS.get("admin_password_salt") else b""
    ADMIN_PASSWORD_HASH = bytes.fromhex(SETTINGS["admin_password_hash"]) if SETTINGS.get("admin_password_hash") else b""

    # A configured ARL makes Deezer the default audio source. A bad or expired ARL must
    # never stop the rest of the program, so any failure just means "no Deezer downloads".
    DEEZER = None
    DEEZER_ERROR = ""
    deezer_arl = str(SETTINGS.get("deezer_arl", ""))
    if deezer_arl:
        try:
            DEEZER = deezer.DeezerClient(deezer_arl)
        except Exception as exc:
            DEEZER_ERROR = str(exc)

    # musicdl sits between Deezer and YouTube as a download fallback and parses the
    # other platforms' URLs. Importing it is slow, so availability is checked lazily
    # on first use rather than here. Its logger insists on a writable XDG directory,
    # which a hardened service account may not have in its home — point it at the
    # state dir instead.
    MUSICDL_ENABLED = bool(SETTINGS.get("musicdl_enabled"))
    if MUSICDL_ENABLED:
        musicdl_source.prepare_environment(STATE_DIR)
    MUSICDL_SOURCES = [
        name.strip()
        for name in str(SETTINGS.get("musicdl_sources", "")).split(",")
        if name.strip()
    ]

    # How much a search brings back, and how hard a download tries before giving up.
    SEARCH_RESULT_LIMIT = config.clamp("search_result_limit", SETTINGS.get("search_result_limit"))
    SEARCH_MUSICDL = bool(SETTINGS.get("search_musicdl")) and MUSICDL_ENABLED
    DOWNLOAD_RETRIES = config.clamp("download_retries", SETTINGS.get("download_retries"))
    DOWNLOAD_RETRY_DELAY = config.clamp("download_retry_delay", SETTINGS.get("download_retry_delay"))
    DOWNLOAD_GAP_SECONDS = config.clamp("download_gap_seconds", SETTINGS.get("download_gap_seconds"))
    RATE_LIMIT_COOLDOWN = config.clamp("rate_limit_cooldown", SETTINGS.get("rate_limit_cooldown"))
    JOB_RETRY_LIMIT = config.clamp("job_retry_limit", SETTINGS.get("job_retry_limit"))
    AUTO_UPDATE_TOOLS = bool(SETTINGS.get("auto_update_tools", True))
    AUTO_UPDATE_INTERVAL_HOURS = config.clamp(
        "auto_update_interval_hours", SETTINGS.get("auto_update_interval_hours")
    )

    app.secret_key = SECRET_KEY or "setup-mode-only"
    # Secure by default. The portable build turns this off when it saves a loopback
    # bind, because a Secure cookie is not sent over plain HTTP.
    app.config["SESSION_COOKIE_SECURE"] = bool(SETTINGS.get("secure_cookies", True))
    signer = URLSafeTimedSerializer(SECRET_KEY or "setup-mode-only", salt="requestcast-result-v1")

    for directory in (STATE_DIR, DOWNLOAD_DIR):
        try:
            directory.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
    if config.is_configured(SETTINGS):
        init_db()


def password_required() -> bool:
    return bool(PASSWORD_HASH)

_rate_lock = threading.Lock()
_rate_events: dict[tuple[str, str], list[float]] = {}


def db_connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, timeout=30)
    con.row_factory = sqlite3.Row
    return con


def init_db() -> None:
    with db_connect() as con:
        con.execute("PRAGMA journal_mode=WAL")
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS jobs (
                id TEXT PRIMARY KEY,
                state TEXT NOT NULL,
                label TEXT NOT NULL,
                detail TEXT NOT NULL DEFAULT '',
                total INTEGER NOT NULL DEFAULT 0,
                completed INTEGER NOT NULL DEFAULT 0,
                error TEXT NOT NULL DEFAULT '',
                payload TEXT NOT NULL,
                created_at INTEGER NOT NULL,
                updated_at INTEGER NOT NULL,
                attempts INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        # Job databases written before retries existed have no attempts column.
        columns = {str(row["name"]) for row in con.execute("PRAGMA table_info(jobs)")}
        if "attempts" not in columns:
            con.execute("ALTER TABLE jobs ADD COLUMN attempts INTEGER NOT NULL DEFAULT 0")
        con.execute(
            "UPDATE jobs SET state='queued', detail='Resuming after service restart' WHERE state='running'"
        )


def rate_allowed(bucket: str, key: str, limit: int, window: int) -> bool:
    now = time.monotonic()
    cutoff = now - window
    with _rate_lock:
        slot = (bucket, key)
        events = [stamp for stamp in _rate_events.get(slot, []) if stamp >= cutoff]
        if len(events) >= limit:
            _rate_events[slot] = events
            return False
        events.append(now)
        _rate_events[slot] = events
        return True


def client_ip() -> str:
    return request.remote_addr or "unknown"


def hash_password(candidate: str, salt: bytes) -> bytes:
    return hashlib.scrypt(
        candidate.encode("utf-8"), salt=salt, n=2**15, r=8, p=1, dklen=32,
        maxmem=64 * 1024 * 1024,
    )


def verify_password(candidate: str) -> bool:
    if not PASSWORD_HASH:
        return True
    return hmac.compare_digest(hash_password(candidate, PASSWORD_SALT), PASSWORD_HASH)


def verify_admin_password(candidate: str) -> bool:
    if not ADMIN_PASSWORD_HASH:
        return False
    return hmac.compare_digest(hash_password(candidate, ADMIN_PASSWORD_SALT), ADMIN_PASSWORD_HASH)


def csrf_token() -> str:
    nonce = session.get("nonce", "")
    return hmac.new(SECRET_KEY.encode(), nonce.encode(), hashlib.sha256).hexdigest()


def require_csrf() -> None:
    supplied = request.form.get("csrf", "")
    if not supplied or not hmac.compare_digest(supplied, csrf_token()):
        abort(400, "The form expired. Please go back and try again.")


@app.before_request
def require_login() -> Any:
    if request.endpoint in {"healthz", "static", "stylesheet"}:
        return None
    # Until setup is finished the only thing the program can usefully show is setup.
    if not config.is_configured(SETTINGS):
        if request.endpoint in {"setup", "setup_tools"}:
            return None
        return redirect(url_for("setup"))
    if request.endpoint == "setup":
        return redirect(url_for("preferences"))
    if request.endpoint == "login":
        return None
    if not password_required():
        session["authenticated"] = True
        session.setdefault("nonce", uuid.uuid4().hex)
    elif not session.get("authenticated"):
        return redirect(url_for("login", next=request.full_path.rstrip("?")))
    if request.endpoint == "admin_login":
        return None
    if request.endpoint in {"preferences", "setup_tools"} and not session.get("admin_authenticated"):
        return redirect(url_for("admin_login", next=request.full_path.rstrip("?")))
    return None


@app.after_request
def security_headers(response):
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; style-src 'self'; img-src 'self' data: https://i.ytimg.com "
        "https://*.ytimg.com https://lh3.googleusercontent.com https://*.dzcdn.net; "
        "frame-src https://www.youtube-nocookie.com; "
        "media-src 'self' https://*.dzcdn.net; form-action 'self'; frame-ancestors 'none'; "
        "base-uri 'none'; script-src 'none'"
    )
    # YouTube embeds require an HTTP Referer for API client identification.
    # This sends only our origin cross-site, without exposing URL paths or queries.
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if request.endpoint not in {"static", "stylesheet"}:
        response.headers["Cache-Control"] = "no-store"
    return response


@app.context_processor
def template_helpers() -> dict[str, Any]:
    return {"csrf_token": csrf_token}


@app.route("/healthz")
def healthz():
    return {"status": "ok"}


@app.route("/assets/style.css")
def stylesheet():
    """Serve the small stylesheet without Waitress's Windows file-wrapper path."""
    css_path = Path(app.static_folder or "") / "style.css"
    css = css_path.read_text(encoding="utf-8")
    response = app.response_class(css, mimetype="text/css")
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if not rate_allowed("login", client_ip(), 6, 300):
            return render_template("login.html", error="Too many attempts. Try again in a few minutes."), 429
        if verify_password(request.form.get("password", "")):
            session.clear()
            session["authenticated"] = True
            session["nonce"] = uuid.uuid4().hex
            session.permanent = True
            destination = request.form.get("next", "")
            if not destination.startswith("/") or destination.startswith("//"):
                destination = url_for("index")
            return redirect(destination)
        return render_template("login.html", error="Incorrect password."), 401
    return render_template("login.html", error=None, next=request.args.get("next", ""))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    setting_admin_password = not ADMIN_PASSWORD_HASH
    if request.method == "POST":
        destination = request.form.get("next", "")
        if not rate_allowed("admin-login", client_ip(), 6, 300):
            return render_template(
                "admin_login.html", error="Too many attempts. Try again later.",
                next=destination, setting_admin_password=setting_admin_password,
            ), 429
        if setting_admin_password:
            require_csrf()
            password = request.form.get("password", "")
            if not password:
                return render_template(
                    "admin_login.html", error="Set an admin password.",
                    next=destination, setting_admin_password=True,
                ), 400
            salt = os.urandom(32)
            merged = {
                **SETTINGS,
                "admin_password_salt": salt.hex(),
                "admin_password_hash": hash_password(password, salt).hex(),
            }
            config.save(merged)
            apply_settings(config.load())
            session["admin_authenticated"] = True
            session.permanent = True
            if not destination.startswith("/") or destination.startswith("//"):
                destination = url_for("preferences")
            return redirect(destination)
        if verify_admin_password(request.form.get("password", "")):
            session["admin_authenticated"] = True
            session.setdefault("nonce", uuid.uuid4().hex)
            session.permanent = True
            if not destination.startswith("/") or destination.startswith("//"):
                destination = url_for("preferences")
            return redirect(destination)
        return render_template(
            "admin_login.html", error="Incorrect admin password.", next=destination,
            setting_admin_password=False,
        ), 401
    return render_template(
        "admin_login.html", error=None, next=request.args.get("next", ""),
        setting_admin_password=setting_admin_password,
    )


@app.post("/logout")
def logout():
    require_csrf()
    session.clear()
    return redirect(url_for("login"))


def api_json(url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    response = requests.get(url, params=params, timeout=HTTP_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, dict) and data.get("error"):
        raise RuntimeError(str(data["error"]))
    return data


def best_thumbnail(thumbnails: list[dict[str, Any]] | None) -> str:
    if not thumbnails:
        return ""
    return str(thumbnails[-1].get("url", ""))


def artists_text(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return ", ".join(str(item.get("name", "")) for item in value if item.get("name"))
    return ""


def format_yt_result(raw: dict[str, Any]) -> dict[str, Any] | None:
    kind = str(raw.get("resultType", ""))
    if kind not in {"song", "video", "album", "playlist", "artist"}:
        return None
    video_id = str(raw.get("videoId") or "")
    browse_id = str(raw.get("browseId") or "")
    title = str(raw.get("title") or raw.get("artist") or "Untitled")
    artist = artists_text(raw.get("artists")) or str(raw.get("artist") or raw.get("author") or "")
    album = raw.get("album") or {}
    album_name = album.get("name", "") if isinstance(album, dict) else str(album)
    duration = raw.get("duration") or ""
    detail_parts = [part for part in (artist, album_name, str(duration)) if part]
    payload = {
        "source": "youtube",
        "kind": kind,
        "id": video_id or browse_id,
        "video_id": video_id,
        "browse_id": browse_id,
        "title": title,
        "artist": artist,
        "album": album_name,
        "duration_seconds": int(raw.get("duration_seconds") or 0),
        "cover": best_thumbnail(raw.get("thumbnails")),
    }
    return {
        **payload,
        "detail": " — ".join(detail_parts),
        "preview_type": "youtube" if video_id else "",
        "preview": video_id,
        "token": signer.dumps(payload),
    }


def search_youtube(query: str, kind: str, limit: int = 0) -> list[dict[str, Any]]:
    filters = {
        "song": "songs",
        "video": "videos",
        "album": "albums",
        "playlist": "playlists",
        "artist": "artists",
    }
    wanted = limit or SEARCH_RESULT_LIMIT
    # ytmusicapi follows continuations until it has this many, so a larger number
    # genuinely brings back more rather than re-cutting the same first page.
    raw_results = YTMusic().search(query, filter=filters.get(kind), limit=wanted)
    results = []
    for raw in raw_results:
        item = format_yt_result(raw)
        if item:
            results.append(item)
    return results[:wanted]


def deezer_cover(raw: dict[str, Any]) -> str:
    for key in ("cover_xl", "cover_big", "picture_xl", "picture_big", "picture_medium"):
        if raw.get(key):
            return str(raw[key])
    album = raw.get("album") or {}
    if isinstance(album, dict):
        return str(album.get("cover_xl") or album.get("cover_big") or album.get("cover_medium") or "")
    return ""


def format_deezer_result(raw: dict[str, Any], kind: str) -> dict[str, Any]:
    if kind == "song":
        artist = str((raw.get("artist") or {}).get("name", ""))
        album = str((raw.get("album") or {}).get("title", ""))
        title = str(raw.get("title", "Untitled"))
        detail = " — ".join(part for part in (artist, album, f"{int(raw.get('duration', 0)) // 60}:{int(raw.get('duration', 0)) % 60:02d}") if part)
        payload = {
            "source": "deezer", "kind": "song", "id": str(raw.get("id", "")),
            "title": title, "artist": artist, "album": album,
            "duration_seconds": int(raw.get("duration") or 0), "track_number": int(raw.get("track_position") or 0),
            "disc_number": int(raw.get("disk_number") or 0), "isrc": str(raw.get("isrc") or ""),
            "cover": deezer_cover(raw),
        }
        preview = str(raw.get("preview") or "")
    elif kind == "artist":
        title = str(raw.get("name", "Unknown artist")); detail = f"{int(raw.get('nb_album') or 0)} albums"
        payload = {"source": "deezer", "kind": kind, "id": str(raw.get("id", "")), "title": title, "artist": title, "cover": deezer_cover(raw)}
        preview = ""
    else:
        title = str(raw.get("title", "Untitled"))
        artist = str((raw.get("artist") or raw.get("creator") or {}).get("name", ""))
        count = int(raw.get("nb_tracks") or 0)
        detail = " — ".join(part for part in (artist, f"{count} tracks" if count else "") if part)
        payload = {"source": "deezer", "kind": kind, "id": str(raw.get("id", "")), "title": title, "artist": artist, "cover": deezer_cover(raw)}
        preview = ""
    return {
        **payload, "detail": detail, "preview_type": "audio" if preview else "",
        "preview": preview, "token": signer.dumps(payload),
    }


def search_deezer_type(query: str, kind: str, limit: int = 0) -> list[dict[str, Any]]:
    """One page's worth of Deezer results, following pages until the limit is reached.

    Deezer caps a single response at 100 rows, so anything larger is paged with the
    ``index`` parameter rather than silently truncated.
    """
    endpoint_kind = "track" if kind == "song" else kind
    wanted = limit or SEARCH_RESULT_LIMIT
    results: list[dict[str, Any]] = []
    index = 0
    while len(results) < wanted:
        page_size = min(DEEZER_PAGE_SIZE, wanted - len(results))
        data = api_json(
            f"https://api.deezer.com/search/{endpoint_kind}",
            {"q": query, "limit": page_size, "index": index, "output": "json"},
        )
        rows = data.get("data") or []
        results.extend(format_deezer_result(item, kind) for item in rows)
        if len(rows) < page_size or not data.get("next"):
            break
        index += len(rows)
    return results[:wanted]


def search_deezer(query: str, kind: str, limit: int = 0) -> list[dict[str, Any]]:
    kinds = [kind] if kind != "all" else ["song", "album", "artist", "playlist"]
    combined: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {pool.submit(search_deezer_type, query, one_kind, limit): one_kind for one_kind in kinds}
        for future in as_completed(futures):
            combined.extend(future.result())
    order = {"song": 0, "video": 1, "album": 2, "artist": 3, "playlist": 4}
    combined.sort(key=lambda item: (order.get(item["kind"], 9), item["title"].casefold()))
    return combined


def format_musicdl_result(song_info: Any) -> dict[str, Any] | None:
    """A musicdl search hit as a downloadable search result, or None if unusable."""
    try:
        if not song_info.with_valid_download_url or not isinstance(song_info.download_url, str):
            return None
    except Exception:
        return None
    payload = musicdl_track_entry(song_info)
    platform = str(payload.get("client") or "").removesuffix("MusicClient")
    detail_parts = [
        part for part in (
            payload["artist"], payload["album"],
            format_duration(payload["duration_seconds"]) if payload["duration_seconds"] else "",
            platform,
        ) if part
    ]
    return {
        **payload, "detail": " — ".join(detail_parts),
        "preview_type": "", "preview": "", "token": signer.dumps(payload),
    }


def search_musicdl(query: str, kind: str, limit: int = 0) -> list[dict[str, Any]]:
    """Search every configured musicdl platform. Only tracks come back from these."""
    if not MUSICDL_ENABLED:
        return []
    if kind not in {"all", "song"}:
        return []
    problem = musicdl_source.availability()
    if problem:
        raise RuntimeError(problem)
    wanted = limit or SEARCH_RESULT_LIMIT
    found = musicdl_source.search(query, MUSICDL_SOURCES, str(STATE_DIR / "musicdl-search"))
    results: list[dict[str, Any]] = []
    # Take from each platform in turn so one chatty source cannot fill the whole page.
    for row in itertools.zip_longest(*found.values()):
        for song_info in row:
            if song_info is None:
                continue
            item = format_musicdl_result(song_info)
            if item:
                results.append(item)
            if len(results) >= wanted:
                return results
    return results


def format_duration(seconds: int) -> str:
    seconds = max(0, int(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}:{seconds:02d}" if hours else f"{minutes}:{seconds:02d}"


def js_runtime_arguments() -> list[str]:
    """Tell yt-dlp where our own Deno is.

    YouTube answers with JavaScript challenges yt-dlp cannot solve by itself, so it hands
    them to Deno. yt-dlp finds a Deno on PATH without help; the copy in RequestCast's
    tools folder has to be pointed at.
    """
    return ["--js-runtimes", f"deno:{DENO}"] if DENO else []


def ytdlp_json(url: str, *, playlist_limit: int | None = None) -> dict[str, Any]:
    command = [
        YTDLP, "--ignore-config", "--no-plugin-dirs",
        "--skip-download", "--no-warnings", "--socket-timeout", "30",
        "--retries", "3", *js_runtime_arguments(),
    ]
    if playlist_limit is None:
        command.append("--no-playlist")
    else:
        command.extend(["--flat-playlist", "--playlist-end", str(playlist_limit)])
    command.extend(["--dump-single-json", url])
    result = subprocess.run(command, capture_output=True, text=True, timeout=120, check=False)
    if result.returncode != 0:
        lines = (result.stderr or result.stdout).strip().splitlines()
        raise RuntimeError(lines[-1] if lines else "Could not read that YouTube URL.")
    try:
        data = json.loads(result.stdout)
    except ValueError as exc:
        raise RuntimeError("YouTube returned invalid metadata for that URL.") from exc
    if not isinstance(data, dict):
        raise RuntimeError("YouTube returned no usable metadata for that URL.")
    return data


def ytdlp_thumbnail(raw: dict[str, Any]) -> str:
    if raw.get("thumbnail"):
        return str(raw["thumbnail"])
    thumbnails = raw.get("thumbnails") or []
    if isinstance(thumbnails, list):
        for thumbnail in reversed(thumbnails):
            if isinstance(thumbnail, dict) and thumbnail.get("url"):
                return str(thumbnail["url"])
    return ""


def youtube_url_result(value: str) -> dict[str, Any]:
    parsed = urlparse(value if "://" in value else f"https://{value}")
    host = (parsed.hostname or "").lower().removeprefix("www.")
    segments = [segment for segment in parsed.path.split("/") if segment]
    query = parse_qs(parsed.query)
    video_id = ""
    playlist_id = ""
    if host == "youtu.be" and segments:
        video_id = segments[0]
    elif host in {"youtube.com", "m.youtube.com", "music.youtube.com", "youtube-nocookie.com"}:
        if parsed.path == "/watch":
            video_id = (query.get("v") or [""])[0]
        elif len(segments) >= 2 and segments[0] in {"shorts", "embed", "live"}:
            video_id = segments[1]
        if not video_id:
            playlist_id = (query.get("list") or [""])[0]
    else:
        raise RuntimeError("Enter a YouTube or Deezer URL.")
    if video_id:
        if not re.fullmatch(r"[A-Za-z0-9_-]{6,20}", video_id):
            raise RuntimeError("That YouTube video URL is not valid.")
        info = ytdlp_json(f"https://www.youtube.com/watch?v={video_id}")
        resolved_id = str(info.get("id") or video_id)
        title = clean_youtube_title(str(info.get("track") or info.get("title") or "Untitled"))
        artist = clean_text(
            str(info.get("artist") or info.get("creator") or info.get("uploader") or info.get("channel") or "")
        ) or "Unknown Artist"
        album = clean_text(str(info.get("album") or ""))
        duration = int(info.get("duration") or 0)
        payload = {
            "source": "youtube", "kind": "video", "id": resolved_id,
            "video_id": resolved_id, "title": title, "artist": artist,
            "album": album, "duration_seconds": duration,
            "cover": ytdlp_thumbnail(info),
        }
        detail = " — ".join(part for part in (artist, album, format_duration(duration)) if part)
        return {
            **payload, "detail": detail, "preview_type": "youtube",
            "preview": resolved_id, "token": signer.dumps(payload),
        }
    if not re.fullmatch(r"[A-Za-z0-9_-]{10,100}", playlist_id):
        raise RuntimeError("That YouTube URL does not contain a valid video or playlist ID.")
    info = ytdlp_json(
        f"https://www.youtube.com/playlist?list={playlist_id}", playlist_limit=1
    )
    artist = clean_text(str(info.get("uploader") or info.get("channel") or ""))
    count = int(info.get("playlist_count") or 0)
    entries = info.get("entries") or []
    cover = ytdlp_thumbnail(info)
    if not cover and entries and isinstance(entries[0], dict):
        cover = ytdlp_thumbnail(entries[0])
    payload = {
        "source": "youtube", "kind": "playlist", "id": playlist_id,
        "browse_id": playlist_id, "title": str(info.get("title") or "YouTube playlist"),
        "artist": artist, "cover": cover, "url_playlist": True,
    }
    detail = " — ".join(part for part in (artist, f"{count} tracks" if count else "") if part)
    return {
        **payload, "detail": detail, "preview_type": "", "preview": "",
        "token": signer.dumps(payload),
    }


def deezer_url_result(value: str) -> dict[str, Any]:
    parsed = urlparse(value if "://" in value else f"https://{value}")
    host = (parsed.hostname or "").lower().removeprefix("www.")
    if host != "deezer.com":
        raise RuntimeError("Enter a direct deezer.com track, album, artist, or playlist URL.")
    segments = [segment for segment in parsed.path.split("/") if segment]
    content_kind = ""
    content_id = ""
    for index, segment in enumerate(segments[:-1]):
        if segment in {"track", "album", "artist", "playlist"} and segments[index + 1].isdigit():
            content_kind = segment
            content_id = segments[index + 1]
            break
    if not content_kind:
        raise RuntimeError("That Deezer URL does not contain a track, album, artist, or playlist ID.")
    raw = api_json(f"https://api.deezer.com/{content_kind}/{content_id}")
    return format_deezer_result(raw, "song" if content_kind == "track" else content_kind)


def musicdl_track_entry(song_info: Any) -> dict[str, Any]:
    """Shape a musicdl SongInfo as a track. The direct-download payload is kept when
    it is storable; otherwise the job re-resolves the track by searching its platform."""
    identifier = str(song_info.get("identifier") or "")
    entry = {
        "source": "musicdl", "kind": "song", "id": identifier,
        "source_id": identifier, "video_id": "",
        "title": clean_text(str(song_info.get("song_name") or "")) or "Untitled",
        "artist": clean_text(str(song_info.get("singers") or "")) or "Unknown Artist",
        "album": clean_text(str(song_info.get("album") or "")),
        "duration_seconds": int(song_info.get("duration_s") or 0),
        "track_number": 0, "disc_number": 0, "year": "", "isrc": "",
        "cover": str(song_info.get("cover_url") or ""),
        "client": str(song_info.get("source") or ""),
    }
    try:
        entry["musicdl"] = musicdl_source.song_info_to_payload(song_info)
    except musicdl_source.MusicdlError:
        pass
    return entry


def musicdl_url_result(value: str, client_name: str) -> dict[str, Any]:
    """Resolve a URL from one of the platforms musicdl supports."""
    if not MUSICDL_ENABLED:
        raise RuntimeError("musicdl support is turned off in Settings.")
    problem = musicdl_source.availability()
    if problem:
        raise RuntimeError(problem)
    song_infos = musicdl_source.parse_url(value, client_name, str(STATE_DIR / "musicdl-parse"), match_key)
    tracks = [musicdl_track_entry(song_info) for song_info in song_infos]
    if not tracks:
        raise RuntimeError("That URL did not produce any tracks RequestCast can use.")
    platform = client_name.removesuffix("MusicClient")
    if len(tracks) == 1:
        track = tracks[0]
        payload = dict(track)
        detail = " — ".join(
            part for part in (track["artist"], track["album"], format_duration(track["duration_seconds"])) if part
        )
        return {
            **payload, "detail": detail, "preview_type": "", "preview": "",
            "token": signer.dumps(payload),
        }
    # A collection is re-parsed when the job runs, so the token stays small.
    cover = next((track["cover"] for track in tracks if track["cover"]), "")
    payload = {
        "source": "musicdl", "kind": "playlist", "id": value, "url": value,
        "client": client_name, "title": f"{platform} collection",
        "artist": "", "cover": cover, "url_playlist": True,
    }
    return {
        **payload, "detail": f"{platform} — {len(tracks)} tracks",
        "preview_type": "", "preview": "", "token": signer.dumps(payload),
    }


def looks_like_media_url(value: str) -> bool:
    lowered = value.strip().lower()
    return "://" in lowered or lowered.startswith(
        ("youtube.com/", "www.youtube.com/", "m.youtube.com/", "music.youtube.com/", "youtu.be/", "deezer.com/", "www.deezer.com/")
        + tuple(f"{host}/" for host in musicdl_source.URL_HOST_PREFIXES)
    )


def resolve_media_url(value: str) -> dict[str, Any]:
    parsed = urlparse(value if "://" in value else f"https://{value}")
    if parsed.scheme not in {"http", "https"} or parsed.username or parsed.password:
        raise RuntimeError("Only normal HTTPS music URLs are supported.")
    host = (parsed.hostname or "").lower().removeprefix("www.")
    if host in {"youtube.com", "m.youtube.com", "music.youtube.com", "youtube-nocookie.com", "youtu.be"}:
        return youtube_url_result(value)
    if host == "deezer.com":
        return deezer_url_result(value)
    client_name = musicdl_source.client_name_for_url(value)
    if client_name:
        return musicdl_url_result(value, client_name)
    raise RuntimeError(
        "That site is not supported. Use a YouTube or Deezer URL, or a URL from one "
        "of the platforms musicdl supports (NetEase, QQ, Kugou, Kuwo, Migu, Spotify, "
        "SoundCloud, TIDAL, Qobuz, Apple Music, and more)."
    )


def recent_jobs(limit: int = 12) -> list[sqlite3.Row]:
    """List recent jobs without reading their payloads.

    A whole-library import stores every indexed entry in its payload, so ``SELECT *``
    here would read tens of megabytes per job just to draw a list of links.
    """
    with db_connect() as con:
        return con.execute(
            "SELECT id, state, label, total, completed FROM jobs "
            "ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()


def search_sources() -> list[tuple[str, str]]:
    """The source choices the search form offers, as (value, label)."""
    choices = [
        ("both", "YouTube and Deezer"),
        ("youtube", "YouTube only"),
        ("deezer", "Deezer only"),
    ]
    if MUSICDL_ENABLED:
        choices.insert(1, ("all", "YouTube, Deezer, and musicdl"))
        choices.append(("musicdl", "musicdl only"))
    return choices


def requested_search_limit(raw: str) -> int:
    """How many results this search asked for, falling back to the saved preference."""
    try:
        wanted = int(str(raw).strip())
    except (TypeError, ValueError):
        return SEARCH_RESULT_LIMIT
    return config.clamp("search_result_limit", wanted)


@app.route("/")
def index():
    query = request.args.get("q", "").strip()
    source = request.args.get("source", "")
    kind = request.args.get("kind", "all")
    limit = requested_search_limit(request.args.get("limit", ""))
    valid_sources = {value for value, _label in search_sources()}
    if source not in valid_sources:
        # musicdl searches are opt-in per install, so the saved preference decides
        # whether the default search includes it.
        source = "all" if (SEARCH_MUSICDL and "all" in valid_sources) else "both"
    if kind not in {"all", "song", "video", "album", "artist", "playlist"}:
        kind = "all"
    results: list[dict[str, Any]] = []
    errors: list[str] = []
    url_input = looks_like_media_url(query)
    if query:
        if len(query) > (2048 if url_input else 160):
            errors.append("URLs must be 2,048 characters or fewer." if url_input else "Search terms must be 160 characters or fewer.")
        elif not rate_allowed("search", client_ip(), 30, 300):
            errors.append("Too many searches. Please wait a few minutes.")
        elif url_input:
            try:
                results.append(resolve_media_url(query))
            except Exception as exc:
                errors.append(f"URL lookup failed: {exc}")
        else:
            tasks = []
            with ThreadPoolExecutor(max_workers=3) as pool:
                if source in {"both", "all", "youtube"}:
                    tasks.append(("YouTube", pool.submit(search_youtube, query, kind, limit)))
                if source in {"both", "all", "deezer"} and kind != "video":
                    tasks.append(("Deezer", pool.submit(search_deezer, query, kind, limit)))
                if source in {"all", "musicdl"} and MUSICDL_ENABLED:
                    tasks.append(("musicdl", pool.submit(search_musicdl, query, kind, limit)))
                for label, future in tasks:
                    try:
                        results.extend(future.result())
                    except Exception as exc:
                        errors.append(f"{label} search failed: {exc}")
    return render_template(
        "index.html", query=query, source=source, kind=kind, results=results,
        errors=errors, jobs=recent_jobs(), url_input=url_input,
        limit=limit, sources=search_sources(),
        # A limit set in the settings file or the environment may not be one of the
        # offered sizes, so it joins the list rather than silently reverting.
        limit_choices=sorted(set(SEARCH_LIMIT_CHOICES) | {limit, SEARCH_RESULT_LIMIT}),
    )


@app.post("/download")
def queue_download():
    require_csrf()
    if not rate_allowed("download", client_ip(), 20, 300):
        abort(429, "Too many downloads were queued. Please wait a few minutes.")
    try:
        payload = signer.loads(request.form.get("token", ""), max_age=86400)
    except (BadSignature, SignatureExpired):
        abort(400, "This search result expired. Search again and retry.")
    if payload.get("source") not in {"youtube", "deezer", "musicdl"}:
        abort(400)
    action = request.form.get("action", "add")
    if action not in {"add", "add_request"}:
        abort(400)
    payload["_request_after_add"] = action == "add_request"
    payload["_request_ip"] = client_ip()
    job_id = queue_job(
        f"{payload.get('title', 'Untitled')} ({payload.get('source')})",
        "Waiting for the downloader",
        payload,
    )
    return redirect(url_for("job_status", job_id=job_id))


def queue_job(label: str, detail: str, payload: dict[str, Any]) -> str:
    """Add one job to the queue and return its identifier."""
    job_id = uuid.uuid4().hex
    now = int(time.time())
    with db_connect() as con:
        con.execute(
            "INSERT INTO jobs (id,state,label,detail,payload,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
            (job_id, "queued", label, detail, json.dumps(payload), now, now),
        )
    return job_id


@app.get("/browse")
def browse_artist():
    """Show one artist's or channel's releases so a person can pick from them."""
    token = request.args.get("token", "")
    try:
        item = signer.loads(token, max_age=86400)
    except (BadSignature, SignatureExpired):
        flash("That result expired. Search again and retry.")
        return redirect(url_for("index"))
    if item.get("kind") not in BROWSABLE_KINDS:
        flash("Only artists and YouTube channels can be opened this way.")
        return redirect(url_for("index"))
    if not rate_allowed("browse", client_ip(), 20, 300):
        abort(429, "Too many artist pages were opened. Please wait a few minutes.")
    try:
        sections = browse_sections(item)
    except Exception as exc:
        flash(f"That artist could not be opened: {exc}")
        return redirect(url_for("index"))
    return render_template("browse.html", item=item, sections=sections, token=token)


@app.post("/browse")
def queue_browse_selection():
    """Queue everything from an artist, or only the parts that were ticked."""
    require_csrf()
    if not rate_allowed("download", client_ip(), 20, 300):
        abort(429, "Too many downloads were queued. Please wait a few minutes.")
    token = request.form.get("token", "")
    try:
        item = signer.loads(token, max_age=86400)
    except (BadSignature, SignatureExpired):
        flash("That result expired. Search again and retry.")
        return redirect(url_for("index"))
    if item.get("kind") not in BROWSABLE_KINDS:
        abort(400)
    scope = request.form.get("scope", "selected")
    if scope not in {"selected", "everything"}:
        abort(400)
    try:
        sections = browse_sections(item)
    except Exception as exc:
        flash(f"That artist could not be opened: {exc}")
        return redirect(url_for("index"))
    if scope == "everything":
        chosen = [entry["token"] for section in sections for entry in section["items"]]
    else:
        chosen = request.form.getlist("select")
    payloads = selection_payloads(sections, chosen)
    if not payloads:
        flash("Tick at least one release, album, or video, or choose Download everything.")
        return redirect(url_for("browse_artist", token=token))

    name = clean_text(str(item.get("artist") or item.get("title") or "")) or "this artist"
    payload = {
        "source": "selection", "kind": "selection", "title": name,
        "items": payloads, "_request_after_add": False, "_request_ip": client_ip(),
    }
    what = "everything" if scope == "everything" else f"{len(payloads)} selected item(s)"
    job_id = queue_job(
        f"{name} ({what})",
        f"Queued {len(payloads)} item(s); waiting for the downloader",
        payload,
    )
    return redirect(url_for("job_status", job_id=job_id))


@app.post("/import")
def queue_file_import():
    require_csrf()
    if not rate_allowed("import", client_ip(), 5, 300):
        abort(429, "Too many files were uploaded. Please wait a few minutes.")
    uploaded = request.files.get("file")
    filename = Path(str(uploaded.filename or "")).name if uploaded else ""
    extension = Path(filename).suffix.lower()
    if not uploaded or not filename:
        flash("Choose a list or playlist file to upload.")
        return redirect(url_for("index"))
    if extension not in IMPORT_EXTENSIONS:
        flash(
            "That file type is not supported. Upload a TXT, XLSX, or PDF list, or an "
            "M3U, M3U8, PLS, XSPF, WPL, ASX, CUE, or foobar2000 FPL playlist."
        )
        return redirect(url_for("index"))
    try:
        entries = parse_import_file(filename, uploaded.stream)
    except Exception as exc:
        flash(f"The file could not be indexed: {exc}")
        return redirect(url_for("index"))
    if not entries:
        flash("No artist, title, track, or supported music URL entries were found in that file.")
        return redirect(url_for("index"))
    choice = request.form.get("artist_tracks", "all")
    if choice not in IMPORT_ARTIST_TRACK_CHOICES:
        choice = "all"
    artist_limit = IMPORT_ARTIST_TRACK_CHOICES[choice]
    payload = {
        "source": "import", "kind": "file", "filename": filename,
        "entries": entries, "artist_limit": artist_limit,
        "_request_after_add": False, "_request_ip": client_ip(),
    }
    job_id = uuid.uuid4().hex
    now = int(time.time())
    per_artist = "every track" if not artist_limit else f"top {artist_limit} tracks"
    label = f"{filename} ({len(entries)} indexed entries, {per_artist} per artist)"
    with db_connect() as con:
        con.execute(
            "INSERT INTO jobs (id,state,label,detail,payload,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
            (
                job_id, "queued", label,
                f"Indexed {len(entries)} unique entries; waiting for the importer",
                json.dumps(payload), now, now,
            ),
        )
    return redirect(url_for("job_status", job_id=job_id))


@app.route("/jobs/<job_id>")
def job_status(job_id: str):
    if not re.fullmatch(r"[0-9a-f]{32}", job_id):
        abort(404)
    with db_connect() as con:
        job = con.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job:
        abort(404)
    return render_template("job.html", job=job)


# A job that is queued or running is still doing work, so clearing leaves it alone.
FINISHED_JOB_STATES = ("completed", "failed")


def clear_jobs(scope: str) -> int:
    """Forget finished downloads, or every download that is not still running."""
    if scope == "all":
        # Deleting a running job's row would leave the worker updating a row that
        # no longer exists, so active work is always kept.
        query = "DELETE FROM jobs WHERE state NOT IN ('queued','running')"
        parameters: tuple[str, ...] = ()
    else:
        query = f"DELETE FROM jobs WHERE state IN ({','.join('?' * len(FINISHED_JOB_STATES))})"
        parameters = FINISHED_JOB_STATES
    with db_connect() as con:
        return int(con.execute(query, parameters).rowcount or 0)


@app.post("/jobs/clear")
def clear_job_history():
    """Remove download history. The downloaded files themselves are never touched."""
    require_csrf()
    scope = request.form.get("scope", "finished")
    if scope not in {"finished", "all"}:
        abort(400)
    removed = clear_jobs(scope)
    if removed:
        flash(
            f"Removed {removed} download{'' if removed == 1 else 's'} from the history. "
            "The downloaded files were not deleted."
        )
    else:
        flash("There was no finished download history to remove.")
    return redirect(url_for("index"))


@app.post("/jobs/<job_id>/delete")
def delete_job(job_id: str):
    """Remove one finished download from the history."""
    require_csrf()
    if not re.fullmatch(r"[0-9a-f]{32}", job_id):
        abort(404)
    with db_connect() as con:
        job = con.execute("SELECT state FROM jobs WHERE id=?", (job_id,)).fetchone()
        if not job:
            abort(404)
        if job["state"] in {"queued", "running"}:
            flash("That download is still working. It can be removed once it finishes.")
            return redirect(url_for("job_status", job_id=job_id))
        con.execute("DELETE FROM jobs WHERE id=?", (job_id,))
    flash("Removed that download from the history. The downloaded file was not deleted.")
    return redirect(url_for("index"))


@app.post("/jobs/<job_id>/retry")
def retry_job(job_id: str):
    """Put a finished download back in the queue, keeping its history entry."""
    require_csrf()
    if not re.fullmatch(r"[0-9a-f]{32}", job_id):
        abort(404)
    if not rate_allowed("download", client_ip(), 20, 300):
        abort(429, "Too many downloads were queued. Please wait a few minutes.")
    with db_connect() as con:
        job = con.execute("SELECT state, label FROM jobs WHERE id=?", (job_id,)).fetchone()
        if not job:
            abort(404)
        if job["state"] in {"queued", "running"}:
            flash("That download is already queued.")
            return redirect(url_for("job_status", job_id=job_id))
        con.execute(
            "UPDATE jobs SET state='queued', detail=?, error='', completed=0, attempts=0, updated_at=? "
            "WHERE id=?",
            ("Queued again by hand", int(time.time()), job_id),
        )
    flash(f"{job['label']} was queued again.")
    return redirect(url_for("job_status", job_id=job_id))


def update_job(job_id: str, **fields: Any) -> None:
    if not fields:
        return
    fields["updated_at"] = int(time.time())
    assignments = ",".join(f"{key}=?" for key in fields)
    values = list(fields.values()) + [job_id]
    with db_connect() as con:
        con.execute(f"UPDATE jobs SET {assignments} WHERE id=?", values)


def claim_job() -> sqlite3.Row | None:
    with db_connect() as con:
        con.execute("BEGIN IMMEDIATE")
        row = con.execute("SELECT * FROM jobs WHERE state='queued' ORDER BY created_at LIMIT 1").fetchone()
        if not row:
            con.commit()
            return None
        con.execute(
            "UPDATE jobs SET state='running', detail='Preparing download', "
            "attempts=attempts+1, updated_at=? WHERE id=?",
            (int(time.time()), row["id"]),
        )
        con.commit()
        return row


def job_attempts(job: sqlite3.Row) -> int:
    """How many times this job has been started, including the run in progress."""
    try:
        return int(job["attempts"] or 0) + 1
    except (IndexError, KeyError, TypeError, ValueError):
        return 1


def clean_text(value: str) -> str:
    return " ".join(str(value or "").replace("\x00", "").split())


def clean_import_value(value: Any) -> str:
    text = clean_text(str(value or "")).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    return text


def import_entry_key(entry: dict[str, str]) -> str:
    if entry.get("url"):
        return f"url:{entry['url'].rstrip('/').casefold()}"
    if entry.get("title"):
        return f"track:{entry.get('artist', '').casefold()}|{entry['title'].casefold()}"
    return f"query:{entry.get('query', '').casefold()}"


def dedupe_import_entries(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    unique: list[dict[str, str]] = []
    seen: set[str] = set()
    for entry in entries:
        key = import_entry_key(entry)
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(entry)
        if len(unique) > MAX_IMPORT_ENTRIES:
            raise RuntimeError(f"Files may contain at most {MAX_IMPORT_ENTRIES:,} unique entries.")
    return unique


def import_entry_from_values(first: Any, second: Any = None) -> dict[str, str] | None:
    left = clean_import_value(first)
    right = clean_import_value(second)
    if not left and not right:
        return None
    if left.casefold() in {"artist", "performer", "band"} and right.casefold() in {
        "title", "song", "track", "track title",
    }:
        return None
    if left.lower().startswith(("http://", "https://")):
        return {"url": left}
    if left and right:
        return {"artist": left, "title": right}
    value = left or right
    if " - " in value:
        artist, title = (clean_import_value(part) for part in value.split(" - ", 1))
        if artist and title:
            return {"artist": artist, "title": title}
    return {"query": value}


def parse_txt_import(stream: Any) -> list[dict[str, str]]:
    raw = stream.read()
    if not isinstance(raw, bytes):
        raw = str(raw).encode("utf-8")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise RuntimeError("The text file is too large.")
    text = decode_import_text(raw, "text file")
    entries = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        values = line.split("\t", 1)
        entry = import_entry_from_values(values[0], values[1] if len(values) > 1 else None)
        if entry:
            entries.append(entry)
        if len(entries) > MAX_RAW_IMPORT_ENTRIES:
            raise RuntimeError(f"Files may contain at most {MAX_IMPORT_ENTRIES:,} entries.")
    return dedupe_import_entries(entries)


def validate_xlsx_archive(stream: Any) -> None:
    stream.seek(0)
    try:
        with zipfile.ZipFile(stream) as archive:
            expanded_size = sum(member.file_size for member in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise RuntimeError("The XLSX file is damaged or is not a real Excel workbook.") from exc
    finally:
        stream.seek(0)
    if expanded_size > 8 * MAX_UPLOAD_BYTES:
        raise RuntimeError("The expanded XLSX workbook is too large.")


def parse_xlsx_import(stream: Any) -> list[dict[str, str]]:
    validate_xlsx_archive(stream)
    workbook = load_workbook(stream, read_only=True, data_only=True)
    entries: list[dict[str, str]] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                first = row[0] if row else None
                second = row[1] if len(row) > 1 else None
                entry = import_entry_from_values(first, second)
                if entry:
                    entries.append(entry)
                if len(entries) > MAX_RAW_IMPORT_ENTRIES:
                    raise RuntimeError(f"Files may contain at most {MAX_IMPORT_ENTRIES:,} entries.")
    finally:
        workbook.close()
    return dedupe_import_entries(entries)


def parse_pdf_import(stream: Any) -> list[dict[str, str]]:
    stream.seek(0)
    reader = PdfReader(stream)
    if len(reader.pages) > MAX_PDF_PAGES:
        raise RuntimeError(f"PDF files may contain at most {MAX_PDF_PAGES:,} pages.")
    entries: list[dict[str, str]] = []
    for page in reader.pages:
        text = page.extract_text(extraction_mode="layout") or ""
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            columns = re.split(r"\s{2,}", line, maxsplit=2)
            entry = import_entry_from_values(columns[0], columns[1] if len(columns) > 1 else None)
            if entry:
                entries.append(entry)
            if len(entries) > MAX_RAW_IMPORT_ENTRIES:
                raise RuntimeError(f"Files may contain at most {MAX_IMPORT_ENTRIES:,} entries.")
    return dedupe_import_entries(entries)


def read_import_bytes(stream: Any, description: str) -> bytes:
    stream.seek(0)
    raw = stream.read()
    if not isinstance(raw, bytes):
        raw = str(raw).encode("utf-8")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise RuntimeError(f"The {description} is too large.")
    return raw


def decode_import_text(raw: bytes, description: str) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"The {description} encoding could not be read.")


def playlist_track_name(location: str) -> str:
    """Turn a playlist's file reference into the name of the music it points at."""
    text = str(location or "").strip().strip('"')
    if re.match(r"(?i)^[a-z][a-z0-9+.-]*://", text) or text.lower().startswith("file:"):
        # Playlists store references as URIs, so the readable name is percent-encoded.
        text = unquote(urlparse(text).path or text).lstrip("/")
    name = re.split(r"[\\/]", text)[-1]
    if Path(name).suffix.lower() in AUDIO_FILE_EXTENSIONS:
        name = Path(name).stem
    name = name.replace("_", " ")
    # Drop the leading track number that ripping software writes into file names.
    name = re.sub(r"^\s*\d{1,3}\s*[-._)]\s+", "", name)
    return clean_import_value(name)


def playlist_entry(location: str = "", title: str = "", artist: str = "") -> dict[str, str] | None:
    """Build one import entry from a playlist's location and its stated metadata."""
    location = str(location or "").strip()
    title = clean_import_value(title)
    artist = clean_import_value(artist)
    if location.lower().startswith(("http://", "https://")) and looks_like_media_url(location):
        return {"url": location}
    if artist and title:
        return {"artist": artist, "title": title}
    if title:
        return import_entry_from_values(title)
    if not location:
        return None
    name = playlist_track_name(location)
    return import_entry_from_values(name) if name else None


def parse_m3u_playlist(text: str) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    pending_artist = ""
    pending_title = ""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.upper().startswith("#EXTINF:"):
            _duration, _, described = line.partition(":")[2].partition(",")
            described = clean_import_value(described)
            if " - " in described:
                pending_artist, pending_title = (
                    clean_import_value(part) for part in described.split(" - ", 1)
                )
            else:
                pending_artist, pending_title = "", described
            continue
        if line.startswith("#"):
            continue
        entry = playlist_entry(line, pending_title, pending_artist)
        pending_artist = pending_title = ""
        if entry:
            entries.append(entry)
    return entries


def parse_pls_playlist(text: str) -> list[dict[str, str]]:
    files: dict[str, str] = {}
    titles: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        match = re.match(r"(?i)^(file|title)(\d+)\s*=\s*(.*)$", line)
        if not match:
            continue
        field, number, value = match.group(1).lower(), match.group(2), match.group(3).strip()
        (files if field == "file" else titles)[number] = value
    entries: list[dict[str, str]] = []
    for number in sorted(set(files) | set(titles), key=lambda value: int(value)):
        entry = playlist_entry(files.get(number, ""), titles.get(number, ""))
        if entry:
            entries.append(entry)
    return entries


def parse_playlist_xml(text: str) -> Any:
    """Parse playlist XML, refusing the entity tricks that make XML a security problem."""
    if re.search(r"<!\s*(DOCTYPE|ENTITY)", text, re.IGNORECASE):
        raise RuntimeError("That playlist declares an XML document type, which is not read.")
    try:
        return ElementTree.fromstring(text)
    except ElementTree.ParseError as exc:
        raise RuntimeError(f"That playlist is not readable XML: {exc}") from exc


def _xml_text(element: Any, *names: str) -> str:
    """Read a child element's text, ignoring the XML namespace it was written with."""
    for child in element.iter():
        tag = str(child.tag or "").rsplit("}", 1)[-1].casefold()
        if tag in names and (child.text or "").strip():
            return str(child.text).strip()
    return ""


def parse_xspf_playlist(text: str) -> list[dict[str, str]]:
    root = parse_playlist_xml(text)
    entries: list[dict[str, str]] = []
    for element in root.iter():
        if str(element.tag or "").rsplit("}", 1)[-1].casefold() != "track":
            continue
        entry = playlist_entry(
            _xml_text(element, "location"),
            _xml_text(element, "title"),
            _xml_text(element, "creator"),
        )
        if entry:
            entries.append(entry)
    return entries


def parse_xml_reference_playlist(text: str) -> list[dict[str, str]]:
    """Read Windows Media Player WPL and ASX playlists, which name files in attributes."""
    root = parse_playlist_xml(text)
    entries: list[dict[str, str]] = []
    for element in root.iter():
        tag = str(element.tag or "").rsplit("}", 1)[-1].casefold()
        if tag not in {"media", "ref", "entry"}:
            continue
        location = ""
        for name, value in element.attrib.items():
            if str(name).rsplit("}", 1)[-1].casefold() in {"src", "href"}:
                location = str(value)
                break
        entry = playlist_entry(location, _xml_text(element, "title"), _xml_text(element, "author"))
        if entry:
            entries.append(entry)
    return entries


def parse_cue_playlist(text: str) -> list[dict[str, str]]:
    album_artist = ""
    entries: list[dict[str, str]] = []
    current_artist = ""
    current_title = ""
    in_track = False

    def flush() -> None:
        nonlocal current_artist, current_title
        if current_title:
            entry = playlist_entry("", current_title, current_artist or album_artist)
            if entry:
                entries.append(entry)
        current_artist = current_title = ""

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if re.match(r"(?i)^TRACK\s+\d+", line):
            flush()
            in_track = True
            continue
        match = re.match(r'(?i)^(PERFORMER|TITLE)\s+"?(.*?)"?\s*$', line)
        if not match:
            continue
        field, value = match.group(1).upper(), clean_import_value(match.group(2))
        if not in_track:
            if field == "PERFORMER":
                album_artist = value
            continue
        if field == "PERFORMER":
            current_artist = value
        else:
            current_title = value
    flush()
    return entries


def parse_fpl_playlist(raw: bytes) -> list[dict[str, str]]:
    """Read a foobar2000 playlist, whose file references sit in a binary string table."""
    entries: list[dict[str, str]] = []
    seen: set[str] = set()
    for chunk in raw.split(b"\x00"):
        try:
            text = chunk.decode("utf-8")
        except UnicodeDecodeError:
            continue
        text = text.strip()
        if len(text) < 4 or len(text) > 2048:
            continue
        lowered = text.casefold()
        is_reference = lowered.startswith(("file://", "http://", "https://")) or (
            Path(lowered).suffix in AUDIO_FILE_EXTENSIONS and re.search(r"[\\/]", text)
        )
        if not is_reference or text in seen:
            continue
        seen.add(text)
        entry = playlist_entry(text)
        if entry:
            entries.append(entry)
    if not entries:
        raise RuntimeError(
            "No track references were found in that foobar2000 playlist. "
            "Export it as M3U8 from foobar2000 and upload that instead."
        )
    return entries


def parse_playlist_import(extension: str, stream: Any) -> list[dict[str, str]]:
    raw = read_import_bytes(stream, "playlist file")
    if extension in {".fpl", ".fb2k-playlist"}:
        return dedupe_import_entries(parse_fpl_playlist(raw))
    text = decode_import_text(raw, "playlist file")
    if extension == ".pls":
        entries = parse_pls_playlist(text)
    elif extension == ".xspf":
        entries = parse_xspf_playlist(text)
    elif extension in {".wpl", ".asx"}:
        entries = parse_xml_reference_playlist(text)
    elif extension == ".cue":
        entries = parse_cue_playlist(text)
    else:
        entries = parse_m3u_playlist(text)
    if not entries:
        raise RuntimeError("No tracks were found in that playlist.")
    return dedupe_import_entries(entries)


def parse_import_file(filename: str, stream: Any) -> list[dict[str, str]]:
    extension = Path(filename).suffix.lower()
    parsers = {".txt": parse_txt_import, ".xlsx": parse_xlsx_import, ".pdf": parse_pdf_import}
    parser = parsers.get(extension)
    if parser:
        return parser(stream)
    if extension in PLAYLIST_EXTENSIONS:
        return parse_playlist_import(extension, stream)
    raise RuntimeError(
        "Only TXT, XLSX, PDF, and playlist files (M3U, M3U8, PLS, XSPF, WPL, ASX, CUE, FPL) "
        "are supported."
    )


def safe_filename(value: str, fallback: str = "Unknown") -> str:
    value = unicodedata.normalize("NFKC", clean_text(value))
    value = re.sub(r"[\\/:*?\"<>|\x00-\x1f]", "-", value).strip(" .-")
    value = re.sub(r"\s+", " ", value)
    return (value[:120].rstrip(" .-") or fallback)


def parse_duration(value: str) -> int:
    parts = [int(part) for part in value.split(":") if part.isdigit()]
    seconds = 0
    for part in parts:
        seconds = seconds * 60 + part
    return seconds


def clean_youtube_title(title: str) -> str:
    title = clean_text(title)
    title = re.sub(
        r"\s*[\[(](official\s+)?(music\s+)?(video|audio|lyric(s)?|visuali[sz]er|hd|hq|4k)[^\])]*[\])]\s*$",
        "", title, flags=re.IGNORECASE,
    )
    return title.strip(" -–—|")


def metadata_from_yt_track(raw: dict[str, Any]) -> dict[str, Any]:
    title = clean_youtube_title(str(raw.get("title") or "Untitled"))
    artist = artists_text(raw.get("artists")) or str(raw.get("author") or "")
    if not artist and " - " in title:
        artist, title = [part.strip() for part in title.split(" - ", 1)]
    album = raw.get("album") or {}
    return {
        "source": "youtube", "source_id": str(raw.get("videoId") or ""),
        "video_id": str(raw.get("videoId") or ""), "title": title,
        "artist": clean_text(artist) or "Unknown Artist",
        "album": clean_text(album.get("name", "") if isinstance(album, dict) else str(album)),
        "duration_seconds": int(raw.get("duration_seconds") or parse_duration(str(raw.get("duration") or ""))),
        "track_number": int(raw.get("trackNumber") or 0), "disc_number": 0,
        "year": str(raw.get("year") or ""), "isrc": "", "cover": best_thumbnail(raw.get("thumbnails")),
    }


def youtube_search_score(candidate: dict[str, Any], wanted: dict[str, Any]) -> float:
    def tokens(value: str) -> set[str]:
        return set(re.findall(r"[a-z0-9]+", unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()))
    desired = tokens(f"{wanted.get('artist', '')} {wanted.get('title', '')}")
    actual = tokens(f"{artists_text(candidate.get('artists'))} {candidate.get('title', '')}")
    overlap = len(desired & actual) / max(1, len(desired))
    wanted_duration = int(wanted.get("duration_seconds") or 0)
    candidate_duration = int(candidate.get("duration_seconds") or parse_duration(str(candidate.get("duration") or "")))
    duration_score = 0.0
    if wanted_duration and candidate_duration:
        delta = abs(wanted_duration - candidate_duration)
        duration_score = max(-1.0, 1.0 - delta / 20.0)
    return overlap * 5 + duration_score


def resolve_deezer_track(item: dict[str, Any], ytm: YTMusic | None = None) -> dict[str, Any]:
    if not item.get("duration_seconds") or not item.get("isrc"):
        raw = api_json(f"https://api.deezer.com/track/{item['id']}")
        item = {
            **item,
            "title": str(raw.get("title") or item.get("title") or "Untitled"),
            "artist": str((raw.get("artist") or {}).get("name") or item.get("artist") or "Unknown Artist"),
            "album": str((raw.get("album") or {}).get("title") or item.get("album") or ""),
            "duration_seconds": int(raw.get("duration") or item.get("duration_seconds") or 0),
            "track_number": int(raw.get("track_position") or item.get("track_number") or 0),
            "disc_number": int(raw.get("disk_number") or item.get("disc_number") or 0),
            "isrc": str(raw.get("isrc") or item.get("isrc") or ""),
            "cover": deezer_cover(raw) or item.get("cover", ""),
            "year": str(raw.get("release_date") or "")[:4],
        }
    query = f"{item.get('artist', '')} {item.get('title', '')}"
    searcher = ytm or YTMusic()
    candidates = searcher.search(query, filter="songs", limit=8)
    candidates = [candidate for candidate in candidates if candidate.get("videoId")]
    if not candidates:
        candidates = searcher.search(query, filter="videos", limit=8)
    if not candidates:
        raise RuntimeError("No matching YouTube audio source was found for this Deezer track.")
    best = max(candidates, key=lambda candidate: youtube_search_score(candidate, item))
    return {
        **item, "source": "deezer", "source_id": str(item["id"]),
        "video_id": str(best["videoId"]), "artist": clean_text(item.get("artist", "")) or "Unknown Artist",
        "title": clean_text(item.get("title", "")) or "Untitled",
    }


def expand_youtube(item: dict[str, Any]) -> list[dict[str, Any]]:
    ytm = YTMusic()
    kind = item["kind"]
    if kind in {"song", "video"}:
        return [{
            **item, "source_id": item.get("video_id") or item.get("id"),
            "video_id": item.get("video_id") or item.get("id"),
            "artist": item.get("artist") or "Unknown Artist", "title": clean_youtube_title(item.get("title", "Untitled")),
        }]
    if kind == "album":
        raw_tracks = ytm.get_album(item["browse_id"])["tracks"]
    elif kind == "playlist":
        if item.get("url_playlist"):
            playlist_id = str(item["browse_id"])
            raw = ytdlp_json(
                f"https://www.youtube.com/playlist?list={playlist_id}",
                playlist_limit=MAX_COLLECTION_TRACKS,
            )
            tracks = []
            for entry in raw.get("entries") or []:
                if not isinstance(entry, dict):
                    continue
                video_id = str(entry.get("id") or "")
                if not re.fullmatch(r"[A-Za-z0-9_-]{6,20}", video_id):
                    continue
                title = clean_youtube_title(str(entry.get("track") or entry.get("title") or "Untitled"))
                artist = clean_text(str(entry.get("artist") or entry.get("uploader") or entry.get("channel") or item.get("artist") or "")) or "Unknown Artist"
                tracks.append({
                    "source": "youtube", "source_id": video_id, "video_id": video_id,
                    "title": title, "artist": artist, "album": clean_text(str(entry.get("album") or "")),
                    "duration_seconds": int(entry.get("duration") or 0), "track_number": 0,
                    "disc_number": 0, "year": "", "isrc": "",
                    "cover": ytdlp_thumbnail(entry),
                })
            return tracks
        playlist_id = item["browse_id"]
        if playlist_id.startswith("VL"):
            playlist_id = playlist_id[2:]
        raw_tracks = ytm.get_playlist(playlist_id, limit=MAX_COLLECTION_TRACKS).get("tracks", [])
    elif kind == "artist":
        artist_data = ytm.get_artist(item["browse_id"])
        raw_tracks = (artist_data.get("songs") or {}).get("results", [])
    else:
        raise RuntimeError("Unsupported YouTube result type.")
    tracks = [metadata_from_yt_track(raw) for raw in raw_tracks if raw.get("videoId")]
    return tracks[:MAX_COLLECTION_TRACKS]


def is_youtube_collection_url(value: str) -> bool:
    parsed = urlparse(value if "://" in value else f"https://{value}")
    host = (parsed.hostname or "").lower().removeprefix("www.")
    if parsed.scheme not in {"http", "https"} or parsed.username or parsed.password:
        return False
    if host not in {"youtube.com", "m.youtube.com", "music.youtube.com"}:
        return False
    segments = [segment for segment in parsed.path.split("/") if segment]
    return bool(segments and (segments[0].startswith("@") or segments[0] in {"channel", "c", "user"}))


def expand_youtube_collection_url(value: str) -> list[dict[str, Any]]:
    raw = ytdlp_json(value, playlist_limit=MAX_COLLECTION_TRACKS)
    fallback_artist = clean_text(str(raw.get("uploader") or raw.get("channel") or "")) or "Unknown Artist"
    tracks = []
    for entry in raw.get("entries") or []:
        if not isinstance(entry, dict):
            continue
        video_id = str(entry.get("id") or "")
        if not re.fullmatch(r"[A-Za-z0-9_-]{6,20}", video_id):
            continue
        title = clean_youtube_title(str(entry.get("track") or entry.get("title") or "Untitled"))
        artist = clean_text(
            str(entry.get("artist") or entry.get("uploader") or entry.get("channel") or fallback_artist)
        ) or fallback_artist
        tracks.append({
            "source": "youtube", "source_id": video_id, "video_id": video_id,
            "title": title, "artist": artist,
            "album": clean_text(str(entry.get("album") or "")),
            "duration_seconds": int(entry.get("duration") or 0), "track_number": 0,
            "disc_number": 0, "year": "", "isrc": "", "cover": ytdlp_thumbnail(entry),
        })
    return tracks


def match_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    text = re.sub(r"[(\[][^)\]]*[)\]]", " ", text)
    text = re.sub(r"\b(feat|ft|featuring|with|and)\b.*$", " ", text)
    return " ".join(re.findall(r"[a-z0-9]+", text))


def deezer_song_item(raw: dict[str, Any]) -> dict[str, Any]:
    formatted = format_deezer_result(raw, "song")
    return {
        key: value for key, value in formatted.items()
        if key not in {"token", "detail", "preview", "preview_type"}
    }


def find_deezer_song(artist: str, title: str) -> dict[str, Any] | None:
    """Return the Deezer track that cleanly matches this artist/title, if one exists."""
    wanted_title = match_key(title)
    wanted_artist = match_key(artist)
    if not wanted_title:
        return None
    queries = []
    if artist:
        queries.append(f'artist:"{artist}" track:"{title}"')
    queries.append(clean_text(f"{artist} {title}"))
    for query in queries:
        try:
            data = api_json("https://api.deezer.com/search", {"q": query, "limit": 10, "output": "json"})
        except Exception:
            continue
        for raw in data.get("data") or []:
            if not raw.get("id"):
                continue
            titles = {match_key(raw.get("title")), match_key(raw.get("title_short"))}
            if wanted_title not in titles:
                continue
            candidate_artist = match_key((raw.get("artist") or {}).get("name"))
            if wanted_artist and not (
                wanted_artist == candidate_artist
                or wanted_artist in candidate_artist
                or candidate_artist in wanted_artist
            ):
                continue
            return raw
    return None


def resolve_import_via_deezer(artist: str, title: str, ytm: YTMusic) -> dict[str, Any] | None:
    try:
        raw = find_deezer_song(artist, title)
        if not raw:
            return None
        return resolve_deezer_track(deezer_song_item(raw), ytm)
    except Exception:
        return None


def deezer_paginate(url: str, params: dict[str, Any] | None, ceiling: int) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    while url and len(output) < ceiling:
        data = api_json(url, params)
        params = None
        output.extend(data.get("data") or [])
        url = str(data.get("next") or "")
    return output[:ceiling]


def deezer_artist_catalog(artist_id: str, wanted: str, ceiling: int) -> list[dict[str, Any]]:
    """Every distinct track Deezer lists for this artist, across their whole discography."""
    albums = deezer_paginate(
        f"https://api.deezer.com/artist/{artist_id}/albums",
        {"limit": 100, "output": "json"},
        MAX_ARTIST_ALBUMS,
    )
    raw_tracks: list[dict[str, Any]] = []
    seen: set[str] = set()
    for album in albums:
        if not album.get("id"):
            continue
        try:
            album_tracks = deezer_paginate(
                f"https://api.deezer.com/album/{album['id']}/tracks",
                {"limit": 100, "output": "json"},
                ceiling,
            )
        except Exception:
            continue
        for raw in album_tracks:
            if not raw.get("id"):
                continue
            # Compilations and "various artists" albums carry other performers; skip those.
            performer = match_key((raw.get("artist") or {}).get("name"))
            if wanted and performer and wanted not in performer and performer not in wanted:
                continue
            key = match_key(raw.get("title_short") or raw.get("title"))
            if not key or key in seen:
                continue
            seen.add(key)
            raw_tracks.append(raw)
            if len(raw_tracks) >= ceiling:
                return raw_tracks
    return raw_tracks


def find_deezer_artist(query: str) -> dict[str, Any] | None:
    wanted = match_key(query)
    if not wanted:
        return None
    try:
        data = api_json("https://api.deezer.com/search/artist", {"q": query, "limit": 10, "output": "json"})
    except Exception:
        return None
    matches = [raw for raw in data.get("data") or [] if raw.get("id") and match_key(raw.get("name")) == wanted]
    if not matches:
        return None
    # Deezer returns unranked namesakes first, so pick the artist people actually listen to.
    return max(matches, key=lambda raw: int(raw.get("nb_fan") or 0))


def resolve_import_artist_via_deezer(query: str, ytm: YTMusic, limit: int = 0) -> list[dict[str, Any]]:
    """Resolve an artist-only entry from Deezer: the full catalogue, or the top `limit` tracks."""
    best = find_deezer_artist(query)
    if not best:
        return []
    ceiling = limit or MAX_IMPORT_TRACKS
    try:
        if limit:
            top = api_json(
                f"https://api.deezer.com/artist/{best['id']}/top",
                {"limit": limit, "output": "json"},
            )
            raw_tracks = [raw for raw in top.get("data") or [] if raw.get("id")][:limit]
        else:
            raw_tracks = deezer_artist_catalog(str(best["id"]), match_key(best.get("name")), ceiling)
    except Exception:
        return []
    tracks = []
    for raw in raw_tracks:
        try:
            tracks.append(resolve_deezer_track(deezer_song_item(raw), ytm))
        except Exception:
            continue
    return tracks


def youtube_artist_tracks(query: str, ytm: YTMusic, limit: int = 0) -> list[dict[str, Any]]:
    artists = [candidate for candidate in ytm.search(query, filter="artists", limit=5) if candidate.get("browseId")]
    if not artists:
        return []
    exact = [
        candidate for candidate in artists
        if clean_text(candidate.get("artist") or candidate.get("title") or "").casefold() == query.casefold()
    ]
    artist_data = ytm.get_artist(str((exact or artists)[0]["browseId"]))
    songs = artist_data.get("songs") or {}
    raw_tracks = list(songs.get("results") or [])
    # The artist page only shows a handful of songs; the linked playlist has the rest.
    if not limit and songs.get("browseId"):
        try:
            playlist_id = str(songs["browseId"]).removeprefix("VL")
            raw_tracks = ytm.get_playlist(playlist_id, limit=None).get("tracks") or raw_tracks
        except Exception:
            pass
    tracks = [metadata_from_yt_track(raw) for raw in raw_tracks if raw.get("videoId")]
    return tracks[:limit] if limit else tracks[:MAX_IMPORT_TRACKS]


def resolve_import_entry(entry: dict[str, str], ytm: YTMusic, artist_limit: int = 0) -> list[dict[str, Any]]:
    if entry.get("url"):
        value = entry["url"]
        if is_youtube_collection_url(value):
            return expand_youtube_collection_url(value)
        item = resolve_media_url(value)
        return expand_youtube(item) if item["source"] == "youtube" else expand_deezer(item)
    if entry.get("title"):
        wanted = {"artist": entry.get("artist", ""), "title": entry["title"]}
        deezer_track = resolve_import_via_deezer(wanted["artist"], wanted["title"], ytm)
        if deezer_track:
            return [deezer_track]
        query = clean_text(f"{wanted['artist']} {wanted['title']}")
        candidates = [candidate for candidate in ytm.search(query, filter="songs", limit=8) if candidate.get("videoId")]
        if not candidates:
            candidates = [candidate for candidate in ytm.search(query, filter="videos", limit=8) if candidate.get("videoId")]
        if not candidates:
            raise RuntimeError("No matching YouTube song was found.")
        track = metadata_from_yt_track(max(candidates, key=lambda candidate: youtube_search_score(candidate, wanted)))
        track["artist"] = clean_import_value(wanted["artist"]) or track["artist"]
        track["title"] = clean_import_value(wanted["title"]) or track["title"]
        return [track]
    query = clean_import_value(entry.get("query", ""))
    deezer_catalog = resolve_import_artist_via_deezer(query, ytm, artist_limit)
    if deezer_catalog:
        return deezer_catalog
    youtube_catalog = youtube_artist_tracks(query, ytm, artist_limit)
    if youtube_catalog:
        return youtube_catalog
    candidates = [candidate for candidate in ytm.search(query, filter="songs", limit=8) if candidate.get("videoId")]
    if not candidates:
        raise RuntimeError("No matching artist or song was found.")
    return [metadata_from_yt_track(candidates[0])]


def track_identity(track: dict[str, Any]) -> str:
    artist = clean_text(track.get("artist", "")).casefold()
    title = clean_text(track.get("title", "")).casefold()
    return f"{artist}|{title}" if artist or title else str(track.get("video_id") or track.get("source_id") or "")


def expand_import(payload: dict[str, Any], job_id: str) -> tuple[list[dict[str, Any]], list[str]]:
    entries = payload.get("entries") or []
    ytm = YTMusic()
    tracks: list[dict[str, Any]] = []
    errors: list[str] = []
    seen: set[str] = set()
    for number, entry in enumerate(entries, 1):
        label = entry.get("url") or clean_text(f"{entry.get('artist', '')} {entry.get('title') or entry.get('query', '')}")
        update_job(job_id, detail=f"Resolving indexed entry {number} of {len(entries)}: {label[:160]}")
        try:
            resolved = resolve_import_entry(entry, ytm, int(payload.get("artist_limit") or 0))
        except Exception as exc:
            errors.append(f"{label}: {exc}")
            continue
        for track in resolved:
            identity = track_identity(track)
            if identity and identity not in seen:
                seen.add(identity)
                tracks.append(track)
                if len(tracks) >= MAX_IMPORT_TRACKS:
                    errors.append(f"The import was limited to {MAX_IMPORT_TRACKS:,} unique tracks.")
                    return tracks, errors
    return tracks, errors


def deezer_tracks(endpoint: str) -> list[dict[str, Any]]:
    url = f"https://api.deezer.com/{endpoint}"
    output: list[dict[str, Any]] = []
    while url and len(output) < MAX_COLLECTION_TRACKS:
        data = api_json(url, {"limit": 100} if "?" not in url else None)
        output.extend(data.get("data", []))
        url = str(data.get("next") or "")
    return output[:MAX_COLLECTION_TRACKS]


def expand_deezer(item: dict[str, Any]) -> list[dict[str, Any]]:
    kind = item["kind"]
    if kind == "song":
        return [resolve_deezer_track(item)]
    if kind == "album":
        raw_tracks = deezer_tracks(f"album/{item['id']}/tracks")
    elif kind == "playlist":
        raw_tracks = deezer_tracks(f"playlist/{item['id']}/tracks")
    elif kind == "artist":
        raw_tracks = deezer_tracks(f"artist/{item['id']}/top?limit=25")
    else:
        raise RuntimeError("Unsupported Deezer result type.")
    tracks = []
    for raw in raw_tracks:
        formatted = format_deezer_result(raw, "song")
        tracks.append(resolve_deezer_track({key: value for key, value in formatted.items() if key not in {"token", "detail", "preview", "preview_type"}}))
    return tracks


def expand_musicdl(item: dict[str, Any]) -> list[dict[str, Any]]:
    """Turn a musicdl result into tracks. Collections are parsed again here because
    the queued payload only kept the URL."""
    if item["kind"] == "song":
        return [dict(item)]
    song_infos = musicdl_source.parse_url(
        str(item.get("url") or item.get("id") or ""),
        str(item.get("client") or ""),
        str(STATE_DIR / "musicdl-parse"),
        match_key,
    )
    return [musicdl_track_entry(song_info) for song_info in song_infos]


def browse_item(result: dict[str, Any], detail: str = "") -> dict[str, Any]:
    """Reduce a search result to the fields the browse page shows and submits."""
    return {
        "token": result["token"],
        "label": clean_text(str(result.get("title") or "")) or "Untitled",
        "detail": detail or str(result.get("detail") or ""),
        "kind": str(result.get("kind") or ""),
        "cover": str(result.get("cover") or ""),
    }


def browse_section(heading: str, note: str, items: list[dict[str, Any]]) -> dict[str, Any] | None:
    return {"heading": heading, "note": note, "items": items} if items else None


def deezer_album_detail(raw: dict[str, Any]) -> str:
    year = str(raw.get("release_date") or "")[:4]
    record_type = str(raw.get("record_type") or "").replace("_", " ").strip().title()
    count = int(raw.get("nb_tracks") or 0)
    return " — ".join(
        part for part in (record_type, year, f"{count} tracks" if count else "") if part
    )


def deezer_browse_sections(item: dict[str, Any]) -> list[dict[str, Any]]:
    """List one Deezer artist's releases and popular tracks."""
    artist_id = str(item.get("id") or "")
    if not artist_id.isdigit():
        raise RuntimeError("That Deezer artist has no usable ID.")
    sections: list[dict[str, Any] | None] = []

    try:
        albums = deezer_paginate(
            f"https://api.deezer.com/artist/{artist_id}/albums",
            {"limit": 100, "output": "json"},
            MAX_BROWSE_ITEMS,
        )
    except Exception:
        albums = []
    releases = [
        browse_item(format_deezer_result(raw, "album"), deezer_album_detail(raw))
        for raw in albums
        if raw.get("id")
    ]
    sections.append(browse_section("Releases", "Whole albums, EPs, and singles.", releases))

    try:
        top = api_json(
            f"https://api.deezer.com/artist/{artist_id}/top",
            {"limit": MAX_BROWSE_ITEMS, "output": "json"},
        )
        raw_tracks = [raw for raw in top.get("data") or [] if raw.get("id")]
    except Exception:
        raw_tracks = []
    tracks = [browse_item(format_deezer_result(raw, "song")) for raw in raw_tracks]
    sections.append(browse_section("Popular tracks", "Individual songs.", tracks))
    return [section for section in sections if section]


def youtube_album_result(raw: dict[str, Any], artist: str) -> dict[str, Any] | None:
    browse_id = str(raw.get("browseId") or "")
    if not browse_id:
        return None
    title = clean_text(str(raw.get("title") or "Untitled"))
    payload = {
        "source": "youtube", "kind": "album", "id": browse_id, "video_id": "",
        "browse_id": browse_id, "title": title, "artist": artist, "album": title,
        "duration_seconds": 0, "cover": best_thumbnail(raw.get("thumbnails")),
    }
    detail = " — ".join(
        part for part in (artist, str(raw.get("type") or ""), str(raw.get("year") or "")) if part
    )
    return {**payload, "detail": detail, "token": signer.dumps(payload)}


def youtube_track_result(raw: dict[str, Any], artist: str) -> dict[str, Any] | None:
    track = metadata_from_yt_track(raw)
    if not track["video_id"]:
        return None
    payload = {
        "source": "youtube", "kind": "video", "id": track["video_id"],
        "video_id": track["video_id"], "title": track["title"],
        "artist": track["artist"] if track["artist"] != "Unknown Artist" else (artist or track["artist"]),
        "album": track["album"], "duration_seconds": track["duration_seconds"],
        "cover": track["cover"],
    }
    detail = " — ".join(
        part for part in (
            payload["artist"], payload["album"], format_duration(payload["duration_seconds"])
        ) if part
    )
    return {**payload, "detail": detail, "token": signer.dumps(payload)}


def youtube_playlist_tracks(ytm: YTMusic, browse_id: str) -> list[dict[str, Any]]:
    """Read a YouTube Music playlist, tolerating playlists that cannot be opened."""
    try:
        playlist = ytm.get_playlist(str(browse_id).removeprefix("VL"), limit=MAX_BROWSE_ITEMS)
    except Exception:
        return []
    return list(playlist.get("tracks") or [])


def youtube_browse_sections(item: dict[str, Any]) -> list[dict[str, Any]]:
    """List one YouTube Music artist's albums, singles, songs, and videos."""
    browse_id = str(item.get("browse_id") or item.get("id") or "")
    if not browse_id:
        raise RuntimeError("That YouTube artist has no usable ID.")
    ytm = YTMusic()
    data = ytm.get_artist(browse_id)
    artist = clean_text(str(data.get("name") or item.get("artist") or item.get("title") or ""))
    sections: list[dict[str, Any] | None] = []

    for heading, key in (("Albums", "albums"), ("Singles and EPs", "singles")):
        group = data.get(key) or {}
        entries = list(group.get("results") or [])
        if group.get("params"):
            # The artist page shows only the first few; this asks for the full shelf.
            try:
                entries = ytm.get_artist_albums(
                    browse_id, str(group["params"]), limit=MAX_BROWSE_ITEMS
                ) or entries
            except Exception:
                pass
        results = [youtube_album_result(raw, artist) for raw in entries[:MAX_BROWSE_ITEMS]]
        sections.append(
            browse_section(heading, "Whole releases.", [browse_item(r) for r in results if r])
        )

    for heading, key in (("Songs", "songs"), ("Videos", "videos")):
        group = data.get(key) or {}
        raw_tracks = list(group.get("results") or [])
        if group.get("browseId"):
            raw_tracks = youtube_playlist_tracks(ytm, str(group["browseId"])) or raw_tracks
        results = [youtube_track_result(raw, artist) for raw in raw_tracks[:MAX_BROWSE_ITEMS]]
        sections.append(
            browse_section(heading, "Individual tracks.", [browse_item(r) for r in results if r])
        )
    return [section for section in sections if section]


def youtube_channel_sections(item: dict[str, Any]) -> list[dict[str, Any]]:
    """List a YouTube channel's published releases, playlists, and videos."""
    from .youtube_collections import collection_tab_url

    collection_url = str(item.get("collection_url") or "")
    if not collection_url:
        raise RuntimeError("That YouTube channel has no usable address.")
    fallback_artist = clean_text(str(item.get("artist") or item.get("title") or ""))
    sections: list[dict[str, Any] | None] = []

    def read_tab(tab: str) -> dict[str, Any]:
        try:
            return ytdlp_json(collection_tab_url(collection_url, tab), playlist_limit=MAX_BROWSE_ITEMS)
        except Exception:
            return {}

    # Each tab is a separate yt-dlp run, so read all three at once rather than in turn.
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {tab: pool.submit(read_tab, tab) for tab in ("releases", "playlists", "videos")}
        tabs = {tab: future.result() for tab, future in futures.items()}

    for heading, tab, note in (
        ("Releases", "releases", "Albums and EPs the channel has published."),
        ("Playlists", "playlists", "Whole playlists from the channel."),
    ):
        raw = tabs[tab]
        items: list[dict[str, Any]] = []
        seen: set[str] = set()
        for entry in raw.get("entries") or []:
            if not isinstance(entry, dict):
                continue
            playlist_id = str(entry.get("id") or "")
            if not re.fullmatch(r"[A-Za-z0-9_-]{10,100}", playlist_id) or playlist_id in seen:
                continue
            seen.add(playlist_id)
            title = clean_text(str(entry.get("title") or "Untitled release"))
            payload = {
                "source": "youtube", "kind": "playlist", "id": playlist_id,
                "browse_id": playlist_id, "title": title, "artist": fallback_artist,
                "cover": ytdlp_thumbnail(entry), "url_playlist": True,
            }
            count = int(entry.get("playlist_count") or 0)
            detail = " — ".join(
                part for part in (fallback_artist, f"{count} tracks" if count else "") if part
            )
            items.append(browse_item({**payload, "detail": detail, "token": signer.dumps(payload)}))
        sections.append(browse_section(heading, note, items))

    raw = tabs["videos"]
    videos: list[dict[str, Any]] = []
    seen_videos: set[str] = set()
    for entry in raw.get("entries") or []:
        if not isinstance(entry, dict):
            continue
        video_id = str(entry.get("id") or "")
        if not re.fullmatch(r"[A-Za-z0-9_-]{6,20}", video_id) or video_id in seen_videos:
            continue
        seen_videos.add(video_id)
        title = clean_youtube_title(str(entry.get("track") or entry.get("title") or "Untitled"))
        artist = clean_text(
            str(entry.get("artist") or entry.get("uploader") or entry.get("channel") or fallback_artist)
        ) or "Unknown Artist"
        duration = int(entry.get("duration") or 0)
        payload = {
            "source": "youtube", "kind": "video", "id": video_id, "video_id": video_id,
            "title": title, "artist": artist, "album": clean_text(str(entry.get("album") or "")),
            "duration_seconds": duration, "cover": ytdlp_thumbnail(entry),
        }
        detail = " — ".join(part for part in (artist, format_duration(duration)) if part)
        videos.append(browse_item({**payload, "detail": detail, "token": signer.dumps(payload)}))
    sections.append(browse_section("Videos", "Individual videos.", videos))
    return [section for section in sections if section]


def browse_sections(item: dict[str, Any]) -> list[dict[str, Any]]:
    """List everything a person can pick from for one artist or channel."""
    kind = str(item.get("kind") or "")
    if kind == "channel":
        return youtube_channel_sections(item)
    if kind != "artist":
        raise RuntimeError("Only artists and YouTube channels can be opened this way.")
    if str(item.get("source") or "") == "deezer":
        return deezer_browse_sections(item)
    return youtube_browse_sections(item)


def selection_payloads(sections: list[dict[str, Any]], tokens: list[str]) -> list[dict[str, Any]]:
    """Turn submitted tokens into payloads, keeping the order the page showed."""
    wanted = set(tokens)
    payloads: list[dict[str, Any]] = []
    for section in sections:
        for entry in section["items"]:
            if entry["token"] not in wanted:
                continue
            try:
                payload = signer.loads(entry["token"], max_age=86400)
            except (BadSignature, SignatureExpired):
                continue
            if payload.get("source") in {"youtube", "deezer"}:
                payloads.append(payload)
    return payloads


def expand_selection(payload: dict[str, Any], job_id: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Expand every release, album, or video a person picked into its tracks."""
    items = payload.get("items") or []
    tracks: list[dict[str, Any]] = []
    errors: list[str] = []
    seen: set[str] = set()
    for number, item in enumerate(items, 1):
        label = clean_text(f"{item.get('artist', '')} {item.get('title', 'Untitled')}") or "Untitled"
        update_job(job_id, detail=f"Reading selection {number} of {len(items)}: {label[:160]}")
        try:
            resolved = expand_youtube(item) if item.get("source") == "youtube" else expand_deezer(item)
        except Exception as exc:
            errors.append(f"{label}: {exc}")
            continue
        for track in resolved:
            identity = track_identity(track)
            if identity and identity not in seen:
                seen.add(identity)
                tracks.append(track)
                if len(tracks) >= MAX_IMPORT_TRACKS:
                    errors.append(f"The selection was limited to {MAX_IMPORT_TRACKS:,} unique tracks.")
                    return tracks, errors
    return tracks, errors


def download_cover(url: str) -> tuple[bytes, str] | None:
    if not url or not url.startswith("https://"):
        return None
    response = requests.get(url, timeout=HTTP_TIMEOUT, stream=True)
    response.raise_for_status()
    content_type = response.headers.get("Content-Type", "image/jpeg").split(";", 1)[0]
    if content_type not in {"image/jpeg", "image/png", "image/webp"}:
        return None
    chunks = []
    size = 0
    for chunk in response.iter_content(65536):
        size += len(chunk)
        if size > 5 * 1024 * 1024:
            return None
        chunks.append(chunk)
    return b"".join(chunks), content_type


def tag_id3(path: Path, track: dict[str, Any], cover: tuple[bytes, str] | None) -> None:
    container = None
    if path.suffix.lower() == ".wav":
        container = WAVE(path)
    elif path.suffix.lower() in {".aif", ".aiff"}:
        container = AIFF(path)
    if container is not None:
        if container.tags is None:
            container.add_tags()
        tags = container.tags
        tags.clear()
    else:
        try:
            tags = ID3(path)
            tags.clear()
        except Exception:
            tags = ID3()
    tags.add(TIT2(encoding=3, text=track["title"]))
    tags.add(TPE1(encoding=3, text=track["artist"]))
    if track.get("album"):
        tags.add(TALB(encoding=3, text=track["album"]))
    if track.get("year"):
        tags.add(TDRC(encoding=3, text=str(track["year"])))
    if track.get("track_number"):
        tags.add(TRCK(encoding=3, text=str(track["track_number"])))
    if track.get("disc_number"):
        tags.add(TPOS(encoding=3, text=str(track["disc_number"])))
    tags.add(COMM(encoding=3, lang="eng", desc="Source", text=f"{track['source'].title()} request import"))
    tags.add(TXXX(encoding=3, desc="SOURCE_ID", text=f"{track['source']}:{track['source_id']}"))
    if track.get("isrc"):
        tags.add(TXXX(encoding=3, desc="ISRC", text=track["isrc"]))
    if cover:
        data, mime = cover
        tags.add(APIC(encoding=3, mime=mime, type=3, desc="Cover", data=data))
    if container is not None:
        container.save()
    else:
        tags.save(path, v2_version=3)


def tag_mp4(path: Path, track: dict[str, Any], cover: tuple[bytes, str] | None) -> None:
    audio = MP4(path)
    audio.clear()
    audio["\xa9nam"] = [track["title"]]
    audio["\xa9ART"] = [track["artist"]]
    if track.get("album"):
        audio["\xa9alb"] = [track["album"]]
    if track.get("year"):
        audio["\xa9day"] = [str(track["year"])]
    if track.get("track_number"):
        audio["trkn"] = [(int(track["track_number"]), 0)]
    if track.get("disc_number"):
        audio["disk"] = [(int(track["disc_number"]), 0)]
    audio["\xa9cmt"] = [f"{track['source'].title()} request import"]
    audio["----:com.apple.iTunes:SOURCE_ID"] = [f"{track['source']}:{track['source_id']}".encode()]
    if track.get("isrc"):
        audio["----:com.apple.iTunes:ISRC"] = [str(track["isrc"]).encode()]
    if cover:
        data, mime = cover
        image_format = MP4Cover.FORMAT_PNG if mime == "image/png" else MP4Cover.FORMAT_JPEG
        if mime in {"image/jpeg", "image/png"}:
            audio["covr"] = [MP4Cover(data, imageformat=image_format)]
    audio.save()


def tag_vorbis_or_flac(path: Path, track: dict[str, Any], cover: tuple[bytes, str] | None) -> None:
    audio = MutagenFile(path)
    if audio is None:
        raise RuntimeError(f"No metadata writer is available for {path.suffix} files.")
    audio.clear()
    values = {
        "title": track["title"],
        "artist": track["artist"],
        "album": track.get("album", ""),
        "date": str(track.get("year") or ""),
        "tracknumber": str(track.get("track_number") or ""),
        "discnumber": str(track.get("disc_number") or ""),
        "isrc": str(track.get("isrc") or ""),
        "comment": f"{track['source'].title()} request import",
        "source_id": f"{track['source']}:{track['source_id']}",
    }
    for key, value in values.items():
        if value:
            audio[key] = [value]
    if cover:
        data, mime = cover
        picture = Picture()
        picture.data = data
        picture.mime = mime
        picture.type = 3
        picture.desc = "Cover"
        if isinstance(audio, FLAC):
            audio.clear_pictures()
            audio.add_picture(picture)
        else:
            audio["metadata_block_picture"] = [base64.b64encode(picture.write()).decode("ascii")]
    audio.save()


def tag_audio(path: Path, track: dict[str, Any]) -> None:
    cover = None
    try:
        cover = download_cover(str(track.get("cover") or ""))
    except Exception:
        pass
    suffix = path.suffix.lower()
    if suffix in {".m4a", ".mp4"}:
        tag_mp4(path, track, cover)
    elif suffix in {".flac", ".oga", ".ogg", ".opus"}:
        tag_vorbis_or_flac(path, track, cover)
    elif suffix in {".mp3", ".mp2", ".aac", ".aif", ".aiff", ".wav"}:
        tag_id3(path, track, cover)
    else:
        audio = MutagenFile(path, easy=True)
        if audio is None:
            raise RuntimeError(f"No metadata writer is available for {suffix} files.")
        audio["title"] = track["title"]
        audio["artist"] = track["artist"]
        if track.get("album"):
            audio["album"] = track["album"]
        audio.save()


def probe_audio(path: Path) -> dict[str, Any]:
    result = subprocess.run(
        [
            FFPROBE, "-v", "error", "-select_streams", "a:0",
            "-show_entries", "stream=codec_name,bit_rate,sample_rate,channels:format=format_name,bit_rate",
            "-of", "json", str(path),
        ],
        capture_output=True, text=True, timeout=60, check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "Downloaded file did not contain valid audio.")
    data = json.loads(result.stdout or "{}")
    streams = data.get("streams") or []
    if not streams or not streams[0].get("codec_name"):
        raise RuntimeError("Downloaded file did not contain a decodable audio stream.")
    return {**streams[0], "format_name": (data.get("format") or {}).get("format_name", "")}


def remux_or_preserve_audio(source: Path, temp_dir: Path) -> Path:
    details = probe_audio(source)
    suffix = source.suffix.lower()
    codec = str(details.get("codec_name") or "").lower()
    if suffix in AZURACAST_AUDIO_EXTENSIONS:
        return source
    remux_suffix = {
        "aac": ".m4a",
        "alac": ".m4a",
        "flac": ".flac",
        "mp2": ".mp2",
        "mp3": ".mp3",
        "opus": ".opus",
        "vorbis": ".ogg",
        "wmav1": ".wma",
        "wmav2": ".wma",
    }.get(codec)
    if remux_suffix:
        remuxed = temp_dir / f"preserved{remux_suffix}"
        command = [
            FFMPEG, "-v", "error", "-y", "-i", str(source), "-map", "0:a:0",
            "-vn", "-c:a", "copy", str(remuxed),
        ]
        result = subprocess.run(command, capture_output=True, text=True, timeout=600, check=False)
        if result.returncode == 0 and remuxed.exists():
            return remuxed
    lossless = temp_dir / "preserved.flac"
    result = subprocess.run(
        [
            FFMPEG, "-v", "error", "-y", "-i", str(source), "-map", "0:a:0",
            "-vn", "-c:a", "flac", "-compression_level", "12", str(lossless),
        ],
        capture_output=True, text=True, timeout=1800, check=False,
    )
    if result.returncode != 0 or not lossless.exists():
        raise RuntimeError(result.stderr.strip() or "Could not convert the unsupported source to lossless FLAC.")
    return lossless


def azuracast_headers() -> dict[str, str]:
    return {"X-API-Key": AZURACAST_API_KEY}


def find_azuracast_media(track: dict[str, Any]) -> dict[str, Any] | None:
    if not AZURACAST_ENABLED:
        return None
    source_id = str(track.get("source_id") or "")
    title = clean_text(track.get("title", ""))
    artist = clean_text(track.get("artist", ""))
    search_terms = [source_id, title]
    for search_term in search_terms:
        if not search_term:
            continue
        response = requests.get(
            station_api("/files"),
            headers=azuracast_headers(),
            params={"searchPhrase": search_term, "rowCount": 50},
            timeout=HTTP_TIMEOUT,
        )
        response.raise_for_status()
        for row in response.json().get("rows", []):
            path = str(row.get("path") or "")
            if not path.startswith(f"{UPLOAD_DIR}/"):
                continue
            if source_id and source_id.casefold() in path.casefold():
                return row
            if (
                clean_text(row.get("title", "")).casefold() == title.casefold()
                and clean_text(row.get("artist", "")).casefold() == artist.casefold()
            ):
                return row
    return None


def upload_to_azuracast(path: Path, filename: str) -> None:
    content_types = {
        ".aac": "audio/aac", ".aif": "audio/aiff", ".aiff": "audio/aiff",
        ".flac": "audio/flac", ".m4a": "audio/mp4", ".mp2": "audio/mpeg",
        ".mp3": "audio/mpeg", ".mp4": "audio/mp4", ".oga": "audio/ogg",
        ".ogg": "audio/ogg", ".opus": "audio/ogg", ".wav": "audio/wav",
        ".wma": "audio/x-ms-wma",
    }
    content_type = content_types.get(path.suffix.lower(), "application/octet-stream")
    with path.open("rb") as audio:
        response = requests.post(
            station_api("/files/upload"),
            headers=azuracast_headers(),
            data={"currentDirectory": UPLOAD_DIR},
            files={"file": (filename, audio, content_type)},
            timeout=(10, 300),
        )
    if not response.ok:
        raise RuntimeError(f"AzuraCast upload failed: {response.text.strip()[:500]}")
    result = response.json()
    if not result.get("success"):
        raise RuntimeError(str(result.get("message") or "AzuraCast rejected the upload."))


def ensure_request_playlist(media: dict[str, Any]) -> None:
    if not AZURACAST_ENABLED or not REQUEST_PLAYLIST_ID:
        return
    playlists = media.get("playlists") or []
    if any(str(playlist.get("id")) == REQUEST_PLAYLIST_ID for playlist in playlists):
        return
    response = requests.put(
        station_api("/files/batch"),
        headers={**azuracast_headers(), "Content-Type": "application/json"},
        json={
            "do": "playlist",
            "files": [str(media["path"])],
            "dirs": [],
            "playlists": [REQUEST_PLAYLIST_ID],
        },
        timeout=HTTP_TIMEOUT,
    )
    if not response.ok:
        raise RuntimeError(f"Could not make the media requestable: {response.text.strip()[:500]}")
    result = response.json()
    if result.get("errors"):
        raise RuntimeError("; ".join(str(error) for error in result["errors"]))


def submit_azuracast_request(unique_id: str, request_ip: str) -> str:
    headers = {
        "User-Agent": "RequestCast/1.0",
        "X-Forwarded-For": request_ip,
        "X-Real-IP": request_ip,
    }
    response = requests.post(
        station_api(f"/request/{unique_id}"),
        headers=headers,
        timeout=HTTP_TIMEOUT,
    )
    try:
        result = response.json()
    except ValueError:
        result = {}
    if not response.ok or not result.get("success"):
        message = result.get("message") or response.text.strip() or f"HTTP {response.status_code}"
        raise RuntimeError(str(message)[:500])
    return str(result.get("message") or "The track was requested successfully.")


# AzuraCast answers that mean the file itself was accepted but the automatic
# request was declined under the station's own request rules. They are not
# download failures: the track is safely in the library and request playlist.
REQUEST_REFUSAL_SIGNS = (
    "played too recently",
    "already requested",
    "not requestable",
)


def request_refused(message: str) -> bool:
    """True when AzuraCast declined the request rather than the request failing.

    These come back when the station's request policy applies — the track was
    played too recently, is already queued, or is not in a requestable playlist.
    The add succeeded; only the automatic request was declined.
    """
    lowered = str(message).lower()
    return any(sign in lowered for sign in REQUEST_REFUSAL_SIGNS)


def deezer_track_id(track: dict[str, Any]) -> str:
    """The Deezer track ID for this track, looking it up by artist/title if needed."""
    if track.get("source") == "deezer":
        direct = str(track.get("source_id") or track.get("id") or "")
        if direct.isdigit():
            return direct
    found = find_deezer_song(str(track.get("artist", "")), str(track.get("title", "")))
    return str(found.get("id") or "") if found else ""


def download_via_deezer(track: dict[str, Any], temp_dir: Path) -> Path:
    """Fetch the track's audio from Deezer: FLAC, then 320 or 128 kbps MP3."""
    if DEEZER is None:
        raise deezer.DeezerError("No Deezer session is configured.")
    track_id = deezer_track_id(track)
    if not track_id:
        raise deezer.DeezerError("This track was not found on Deezer.")
    downloaded, _quality = DEEZER.download(track_id, temp_dir)
    return downloaded


def download_via_musicdl(track: dict[str, Any], temp_dir: Path) -> Path:
    """Fetch the track's audio through musicdl, the fallback between Deezer and YouTube."""
    if not MUSICDL_ENABLED:
        raise musicdl_source.MusicdlError("musicdl support is turned off.")
    stored = track.get("musicdl")
    if isinstance(stored, dict):
        return musicdl_source.download_payload(stored, temp_dir)
    # Tracks whose download URL could not be stored (HLS streams, short-lived
    # links) are re-resolved by searching the platform they came from.
    platform = str(track.get("client") or "")
    sources = [platform] if platform else MUSICDL_SOURCES
    return musicdl_source.search_and_download(
        str(track.get("artist", "")), str(track.get("title", "")),
        int(track.get("duration_seconds") or 0), sources, temp_dir, match_key,
    )


def looks_rate_limited(message: str) -> bool:
    """True when a failure reads like the site pushing back rather than a dead track.

    Bulk runs — a channel, a discography, a playlist import — hit this in batches: a
    stretch of tracks fails with 403 or "video unavailable" and the very same tracks
    download fine once the queue slows down.
    """
    lowered = str(message).lower()
    return any(sign in lowered for sign in RATE_LIMIT_SIGNS)


def describe_download_error(message: str) -> str:
    """Say what a raw downloader error means, so the history is worth reading."""
    text = " ".join(str(message).split())
    lowered = text.lower()
    if "javascript runtime" in lowered or "js runtime" in lowered or "jsi" in lowered:
        return f"{text} (YouTube needs a JavaScript runtime. Open Preferences and install the download tools, which include Deno.)"
    if not DENO and ("403" in lowered or "forbidden" in lowered or "nsig" in lowered or "signature" in lowered):
        return f"{text} (Deno is not installed. YouTube's JavaScript challenges cannot be answered without it, which shows up as 403 and missing formats. Install the download tools from Preferences.)"
    if "403" in lowered or "forbidden" in lowered:
        return f"{text} (YouTube refused the download. This is usually rate limiting during a bulk run: it clears on a retry, a longer gap between downloads, or a newer yt-dlp.)"
    if "sign in to confirm" in lowered or "not a bot" in lowered:
        return f"{text} (YouTube asked the downloader to prove it is not a bot, which is rate limiting. A longer gap between downloads clears it.)"
    if "429" in lowered or "too many requests" in lowered:
        return f"{text} (Too many requests were made too quickly. Raise the gap between downloads in Preferences.)"
    if "video unavailable" in lowered or "unavailable" in lowered:
        return f"{text} (The video did not load. During a bulk import this is usually temporary rather than a missing track.)"
    return text


def retry_delay_for(retry: int) -> int:
    """Seconds to wait before retry number ``retry``, doubling each time."""
    if DOWNLOAD_RETRY_DELAY <= 0:
        return 0
    ceiling = max(RATE_LIMIT_COOLDOWN, DOWNLOAD_RETRY_DELAY)
    return int(min(DOWNLOAD_RETRY_DELAY * (2 ** max(0, retry - 1)), ceiling))


def pause_job(job_id: str, seconds: int, reason: str) -> None:
    """Wait, saying on the status page why the job is waiting."""
    if seconds <= 0:
        return
    update_job(job_id, detail=f"{reason} Waiting {seconds} seconds.")
    time.sleep(seconds)


def pace_downloads() -> None:
    """Leave a gap between tracks so a long queue does not trip the site's rate limits."""
    if DOWNLOAD_GAP_SECONDS > 0:
        time.sleep(DOWNLOAD_GAP_SECONDS)


def cool_down(job_id: str, multiplier: int = 1, reason: str = "") -> None:
    """Stop downloading for a while after a run of rate-limited failures."""
    seconds = int(min(RATE_LIMIT_COOLDOWN * max(1, multiplier), 3600))
    pause_job(
        job_id, seconds,
        reason or "Several downloads in a row were refused, which means the site is rate limiting.",
    )


def download_track(
    track: dict[str, Any], job_id: str, label: str, prefix: str = "", attempts: int = 0
) -> tuple[str, dict[str, Any]]:
    """Download one track, retrying as configured. Raises the last error if all fail."""
    attempts = attempts or DOWNLOAD_RETRIES + 1
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            status, _path, media = download_one(track, attempt - 1)
            return status, media
        except Exception as exc:
            last_error = exc
            if attempt >= attempts:
                break
            pause_job(
                job_id, retry_delay_for(attempt),
                f"{prefix}{label} failed on attempt {attempt} of {attempts}: {exc}. Retrying.",
            )
    raise last_error if last_error else RuntimeError("The download failed for an unknown reason.")


def run_downloads(tracks: list[dict[str, Any]], job_id: str) -> tuple[int, str, list[str]]:
    """Download every track, pacing the queue and retrying what the site refuses.

    Returns the number that succeeded, the first AzuraCast request ID seen, and one
    message per track that never made it.
    """
    completed = 0
    first_request_id = ""
    failures: list[tuple[dict[str, Any], str, str]] = []
    streak = 0
    cooldowns = 0
    total = len(tracks)
    for number, track in enumerate(tracks, 1):
        label = f"{track.get('artist', 'Unknown Artist')} — {track.get('title', 'Untitled')}"
        if number > 1:
            pace_downloads()
        update_job(job_id, detail=f"Downloading {number} of {total}: {label}")
        try:
            status, media = download_track(track, job_id, label, f"Track {number} of {total}: ")
        except Exception as exc:
            failures.append((track, label, str(exc)))
            if looks_rate_limited(str(exc)):
                streak += 1
                if streak >= RATE_LIMIT_STREAK and number < total:
                    cooldowns += 1
                    cool_down(job_id, cooldowns)
                    streak = 0
            else:
                streak = 0
            continue
        streak = 0
        completed += 1
        if not first_request_id:
            first_request_id = str(media.get("unique_id") or "")
        update_job(job_id, completed=completed, detail=f"{status.title()}: {label}")

    # Tracks refused during a bulk run usually download fine a few minutes later, so
    # everything that failed gets one more pass once the queue has been quiet.
    if failures and DOWNLOAD_RETRIES > 0 and any(looks_rate_limited(error) for _t, _l, error in failures):
        cool_down(
            job_id, cooldowns + 1,
            f"{len(failures)} track(s) were refused. Waiting before one final pass over them.",
        )
        remaining: list[tuple[dict[str, Any], str, str]] = []
        for number, (track, label, _error) in enumerate(failures, 1):
            if number > 1:
                pace_downloads()
            update_job(job_id, detail=f"Final pass {number} of {len(failures)}: {label}")
            try:
                status, media = download_track(track, job_id, label, "Final pass: ", attempts=1)
            except Exception as exc:
                remaining.append((track, label, str(exc)))
                continue
            completed += 1
            if not first_request_id:
                first_request_id = str(media.get("unique_id") or "")
            update_job(job_id, completed=completed, detail=f"{status.title()} on the final pass: {label}")
        failures = remaining

    errors = [f"{label}: {describe_download_error(error)}" for _track, label, error in failures]
    return completed, first_request_id, errors


def download_one(track: dict[str, Any], attempt: int = 0) -> tuple[str, Path, dict[str, Any]]:
    video_id = str(track.get("video_id") or "")
    source_id = safe_filename(str(track.get("source_id") or video_id), video_id)
    stem = f"{safe_filename(track.get('artist', 'Unknown Artist'))} - {safe_filename(track.get('title', 'Untitled'))} [{track['source']}-{source_id}]"
    existing_media = find_azuracast_media(track)
    if existing_media:
        ensure_request_playlist(existing_media)
        existing_path = MEDIA_DIR / Path(str(existing_media["path"])).name
        return "already existed", existing_path, existing_media
    local_path: Path | None = None
    with tempfile.TemporaryDirectory(prefix="requestcast-", dir=STATE_DIR) as temp_name:
        temp_dir = Path(temp_name)
        prepared: Path | None = None
        # A signed-in Deezer account is the default source. Anything it cannot
        # supply falls through to musicdl and then to the YouTube path below.
        if track.get("source") != "musicdl" and DEEZER is not None:
            try:
                prepared = download_via_deezer(track, temp_dir)
            except Exception:
                prepared = None
        if prepared is None and MUSICDL_ENABLED:
            try:
                prepared = download_via_musicdl(track, temp_dir)
            except Exception:
                prepared = None
        if prepared is None:
            if not re.fullmatch(r"[A-Za-z0-9_-]{6,20}", video_id):
                raise RuntimeError("The selected result has no valid YouTube audio source.")
            if not YTDLP or not DENO:
                # The tools install themselves in the background, so a download that
                # arrives first waits for that rather than failing outright. A missing
                # Deno is not fatal on its own: yt-dlp still works, with fewer formats.
                install_error = ""
                try:
                    ensure_tools_installed()
                except Exception as exc:
                    install_error = str(exc)
                if not YTDLP:
                    raise RuntimeError(
                        f"yt-dlp is missing and could not be installed: {install_error}."
                        if install_error else
                        "yt-dlp was not found. Open Preferences and install the download "
                        "tools, or put yt-dlp on your PATH."
                    )
            command = [
                YTDLP, "--ignore-config", "--no-plugin-dirs",
                "--no-playlist", "--no-progress", "--no-warnings", "--socket-timeout", "30",
                "--retries", str(max(5, DOWNLOAD_RETRIES + 1)),
                "--fragment-retries", str(max(5, DOWNLOAD_RETRIES + 1)),
                "--retry-sleep", f"exp={max(1, DOWNLOAD_RETRY_DELAY // 4)}:{max(4, DOWNLOAD_RETRY_DELAY)}",
                "--sleep-requests", str(max(1, DOWNLOAD_GAP_SECONDS)),
                "--max-filesize", "500M", "--match-filter", "!is_live & duration < 7200",
                "-f", "bestaudio/best", *js_runtime_arguments(),
            ]
            # A 403 or "unavailable" from one YouTube client is often fine from another,
            # so each retry asks as a different one.
            player_clients = YTDLP_PLAYER_CLIENTS[attempt % len(YTDLP_PLAYER_CLIENTS)]
            if player_clients:
                command += ["--extractor-args", f"youtube:player_client={player_clients}"]
            command += [
                "-o", str(temp_dir / "%(id)s.%(ext)s"),
                f"https://www.youtube.com/watch?v={video_id}",
            ]
            result = subprocess.run(command, capture_output=True, text=True, timeout=1800, check=False)
            if result.returncode != 0:
                error = (result.stderr or result.stdout).strip().splitlines()
                raise RuntimeError(error[-1] if error else "yt-dlp failed")
            candidates = [path for path in temp_dir.iterdir() if path.is_file()]
            if not candidates:
                raise RuntimeError("yt-dlp completed but did not produce an audio file.")
            downloaded = max(candidates, key=lambda path: path.stat().st_size)
            prepared = remux_or_preserve_audio(downloaded, temp_dir)
        track["artist"] = clean_text(track.get("artist", "")) or "Unknown Artist"
        track["title"] = clean_youtube_title(track.get("title", "")) or "Untitled"
        tag_audio(prepared, track)
        probe_audio(prepared)
        permissions.make_readable(prepared)
        destination_name = f"{stem}{prepared.suffix.lower()}"
        if AZURACAST_ENABLED:
            upload_to_azuracast(prepared, destination_name)
        else:
            # Local downloader mode: the download folder is the final destination.
            DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
            permissions.make_readable(DOWNLOAD_DIR)
            local_path = DOWNLOAD_DIR / destination_name
            shutil.move(str(prepared), local_path)
            # Moving keeps the temporary folder's private permissions, so set them
            # again once the file is in its final home.
            permissions.make_readable(local_path)
    if not AZURACAST_ENABLED:
        assert local_path is not None
        return "downloaded", local_path, {"path": str(local_path), "unique_id": ""}
    for _ in range(10):
        media = find_azuracast_media(track)
        if media:
            ensure_request_playlist(media)
            media_path = MEDIA_DIR / Path(str(media["path"])).name
            return "downloaded", media_path, media
        time.sleep(1)
    raise RuntimeError("AzuraCast accepted the file but did not return its media record.")


def process_job(job: sqlite3.Row) -> None:
    job_id = job["id"]
    payload = json.loads(job["payload"])
    errors = []
    if payload["source"] == "import":
        tracks, resolution_errors = expand_import(payload, job_id)
        errors.extend(resolution_errors)
    elif payload["source"] == "selection":
        tracks, resolution_errors = expand_selection(payload, job_id)
        errors.extend(resolution_errors)
    else:
        if payload["source"] == "youtube":
            tracks = expand_youtube(payload)
        elif payload["source"] == "musicdl":
            tracks = expand_musicdl(payload)
        else:
            tracks = expand_deezer(payload)
    if not tracks:
        raise RuntimeError("\n".join(errors) or "This result did not contain any downloadable tracks.")
    update_job(job_id, total=len(tracks), detail=f"Found {len(tracks)} track(s)")
    completed, first_request_id, download_errors = run_downloads(tracks, job_id)
    errors.extend(download_errors)
    if completed == 0:
        raise RuntimeError("; ".join(errors) or "No tracks were downloaded.")
    if AZURACAST_ENABLED:
        detail = f"Finished: {completed} of {len(tracks)} track(s) are ready to request."
    else:
        detail = f"Finished: {completed} of {len(tracks)} track(s) saved to {DOWNLOAD_DIR}."
    if payload.get("_request_after_add") and AZURACAST_ENABLED:
        if not first_request_id:
            errors.append("The files were added, but AzuraCast did not return a request ID.")
        else:
            update_job(job_id, detail="Added successfully; submitting the request")
            try:
                request_message = submit_azuracast_request(
                    first_request_id, str(payload.get("_request_ip") or "127.0.0.1")
                )
                detail += f" {request_message}"
            except Exception as exc:
                message = str(exc)
                if request_refused(message):
                    # The track is safely in the library and request playlist;
                    # AzuraCast only declined to auto-queue the request under its
                    # own rules. Say so clearly instead of marking the job failed,
                    # which reads as though the download never happened.
                    detail += f" Added successfully; AzuraCast did not accept the request: {message}"
                else:
                    errors.append(f"Added successfully, but request submission failed: {message}")
    if errors:
        detail += f" {len(errors)} failed."
    update_job(job_id, state="completed", detail=detail, error="\n".join(errors)[:8000])


def fail_or_requeue(job: sqlite3.Row, exc: Exception) -> None:
    """Put a failed job back in the queue while it has retries left, or record the failure."""
    attempts = job_attempts(job)
    message = describe_download_error(str(exc))
    if attempts <= JOB_RETRY_LIMIT:
        update_job(
            job["id"], state="queued",
            detail=f"Attempt {attempts} failed. Retrying (attempt {attempts + 1} of {JOB_RETRY_LIMIT + 1}).",
            error=message[:8000],
        )
        # Let the site settle before the queue picks this job up again.
        pause_job(job["id"], retry_delay_for(attempts), f"Attempt {attempts} failed: {message}")
        return
    update_job(job["id"], state="failed", detail="Download failed", error=message[:8000])


def worker_loop() -> None:
    while True:
        if not config.is_configured(SETTINGS):
            time.sleep(2)
            continue
        try:
            job = claim_job()
        except sqlite3.Error:
            time.sleep(2)
            continue
        if not job:
            time.sleep(2)
            continue
        try:
            process_job(job)
        except Exception as exc:
            fail_or_requeue(job, exc)


# Which setting records where each self-updating tool lives.
TOOL_PATH_SETTINGS = {"yt-dlp": "ytdlp_path", "deno": "deno_path"}
_tool_install_lock = threading.Lock()


def remember_settings(merged: dict[str, Any]) -> None:
    """Use these settings now, and write them down if the settings file can be written.

    Settings do not always live somewhere writable: a service confined to a handful of
    paths, a portable copy on read-only media. By the time a tool path is recorded the
    tool itself is already installed, and ``find_tool`` looks in the tools folder before
    PATH, so an unwritable settings file costs nothing but the note of where it went. It
    must never undo the install, fail the download that asked for it, or stop the
    updater from ever running again.
    """
    if config.is_configured(merged):
        try:
            config.save(merged)
        except OSError:
            pass
    apply_settings(merged)


def ensure_tools_installed(progress: Any = None) -> dict[str, str]:
    """Fetch any missing download tool without waiting to be asked.

    yt-dlp, ffmpeg, and Deno are what make a download work at all, so RequestCast
    installs them itself rather than leaving someone to discover a broken download and
    a button. Anything already present is left alone.
    """
    if WORKER_DISABLED or not tools.installable_tools(SETTINGS):
        return {}
    with _tool_install_lock:
        if not tools.installable_tools(SETTINGS):
            return {}
        updates = tools.install_missing(SETTINGS, progress)
        if updates:
            remember_settings({**SETTINGS, **updates})
        return updates


def save_tool_paths(results: list[dict[str, Any]]) -> None:
    """Remember where an update put a tool, so the next run uses the new copy."""
    updates = {
        TOOL_PATH_SETTINGS[str(item.get("name"))]: str(item.get("path"))
        for item in results
        if item.get("path") and str(item.get("name")) in TOOL_PATH_SETTINGS
    }
    if not updates:
        return
    remember_settings({**SETTINGS, **updates})


def tool_update_loop() -> None:
    """Keep yt-dlp and musicdl current in the background.

    Both read sites that change under them, and an out-of-date yt-dlp is the most common
    cause of downloads that fail with 403 while the same track plays fine in a browser.
    """
    while True:
        if not config.is_configured(SETTINGS):
            time.sleep(30)
            continue
        # Missing tools are installed whether or not automatic updating is on: without
        # them nothing downloads at all.
        try:
            installed = ensure_tools_installed()
        except Exception:
            installed = {}
        if installed:
            tools.record_update_check(
                STATE_DIR,
                [{"name": name, "status": "updated", "message": f"Installed {path}."}
                 for name, path in installed.items()],
            )
        if not AUTO_UPDATE_TOOLS:
            time.sleep(600)
            continue
        due = tools.last_update_check(STATE_DIR) + AUTO_UPDATE_INTERVAL_HOURS * 3600
        if time.time() < due:
            time.sleep(min(3600, max(60, due - time.time())))
            continue
        try:
            results = tools.update_all(SETTINGS)
        except Exception:
            # A failed check must never stop the program; the next one tries again.
            results = []
        tools.record_update_check(STATE_DIR, results)
        save_tool_paths(results)


def tool_update_context() -> dict[str, Any]:
    """What the preferences page says about the tools and when they were last checked."""
    versions: list[tuple[str, str]] = []
    ytdlp_path = tools.find_tool("yt-dlp", SETTINGS.get("ytdlp_path", ""))
    versions.append(("yt-dlp", tools.ytdlp_version(ytdlp_path) or "not installed"))
    deno_path = tools.find_tool("deno", SETTINGS.get("deno_path", ""))
    versions.append((
        "Deno (JavaScript runtime for YouTube)",
        tools.deno_version(deno_path) or "not installed",
    ))
    versions.append(("musicdl", tools.installed_package_version("musicdl") or "not installed"))
    checked = tools.last_update_check(STATE_DIR)
    return {
        "tool_versions": versions,
        "last_update_check": time.strftime("%Y-%m-%d %H:%M", time.localtime(checked)) if checked else "",
    }


def settings_page():
    """Handle first-run setup and later admin preferences."""
    settings = dict(SETTINGS)
    first_run = not config.is_configured(settings)
    if request.method == "POST":
        if not first_run:
            require_csrf()
        submitted = {
            "download_dir": request.form.get("download_dir", "").strip(),
            "azuracast_enabled": bool(request.form.get("azuracast_enabled")),
            "azuracast_api_base": request.form.get("azuracast_api_base", "").strip().rstrip("/"),
            "azuracast_api_key": request.form.get("azuracast_api_key", "").strip(),
            "azuracast_station_id": request.form.get("azuracast_station_id", "").strip() or config.DEFAULT_STATION_ID,
            "azuracast_request_playlist_id": request.form.get("azuracast_request_playlist_id", "").strip(),
            "azuracast_media_dir": request.form.get("azuracast_media_dir", "").strip(),
            "azuracast_upload_dir": request.form.get("azuracast_upload_dir", "").strip() or config.DEFAULT_UPLOAD_DIRECTORY,
            "bind_host": request.form.get("bind_host", "").strip() or config.DEFAULT_BIND_HOST,
            "bind_port": request.form.get("bind_port", "").strip() or config.DEFAULT_BIND_PORT,
            "deezer_arl": request.form.get("deezer_arl", "").strip(),
            "musicdl_enabled": bool(request.form.get("musicdl_enabled")),
            "musicdl_sources": request.form.get("musicdl_sources", "").strip() or musicdl_source.DEFAULT_SOURCES,
            "search_musicdl": bool(request.form.get("search_musicdl")),
            "auto_update_tools": bool(request.form.get("auto_update_tools")),
            "diagnostics_enabled": bool(request.form.get("diagnostics_enabled")),
        }
        # Numbers are clamped rather than rejected, so a typo cannot lock anyone out of
        # their own settings page.
        for key in (
            "search_result_limit", "download_retries", "download_retry_delay",
            "download_gap_seconds", "rate_limit_cooldown", "job_retry_limit",
            "auto_update_interval_hours",
        ):
            submitted[key] = config.clamp(
                key, request.form.get(key, "").strip() or settings.get(key, config.FIELDS[key])
            )
        # Serving plain HTTP on this machine means the session cookie cannot be Secure.
        submitted["secure_cookies"] = submitted["bind_host"] not in config.LOOPBACK_HOSTS
        password = request.form.get("password", "")
        admin_password = request.form.get("admin_password", "")
        problems = validate_setup(
            submitted,
            password,
            clear_password=bool(request.form.get("clear_password")),
            require_passwords=first_run,
            admin_password=admin_password,
        )
        if problems:
            for problem in problems:
                flash(problem)
            return render_template(
                "setup.html", settings={**settings, **submitted},
                first_run=first_run, missing_tools=tools.missing_tools(settings),
                can_auto_install=tools.can_auto_install(), config_path=str(config.config_path()),
                form_endpoint="setup" if first_run else "preferences",
                **tool_update_context(),
            ), 400
        merged = {**settings, **submitted}
        if password:
            salt = os.urandom(32)
            merged["password_salt"] = salt.hex()
            merged["password_hash"] = hash_password(password, salt).hex()
        elif request.form.get("clear_password"):
            merged["password_salt"] = ""
            merged["password_hash"] = ""
        if admin_password:
            admin_salt = os.urandom(32)
            merged["admin_password_salt"] = admin_salt.hex()
            merged["admin_password_hash"] = hash_password(admin_password, admin_salt).hex()
        if not merged.get("secret_key"):
            merged["secret_key"] = config.new_secret_key()
        merged["state_dir"] = merged.get("state_dir") or str(config.default_state_dir())
        config.save(merged)
        apply_settings(config.load())
        session["authenticated"] = True
        session["admin_authenticated"] = True
        session.setdefault("nonce", uuid.uuid4().hex)
        flash("Setup saved." if first_run else "Preferences saved.")
        return redirect(url_for("index" if first_run else "preferences"))
    return render_template(
        "setup.html", settings=settings, first_run=first_run,
        missing_tools=tools.missing_tools(settings),
        can_auto_install=tools.can_auto_install(), config_path=str(config.config_path()),
        form_endpoint="setup" if first_run else "preferences",
        **tool_update_context(),
    )


@app.route("/setup", methods=["GET", "POST"])
def setup():
    return settings_page()


@app.route("/preferences", methods=["GET", "POST"])
def preferences():
    return settings_page()


def validate_setup(
    submitted: dict[str, Any], password: str, *, clear_password: bool = False,
    require_passwords: bool = False, admin_password: str = "",
) -> list[str]:
    problems: list[str] = []
    download_dir = submitted.get("download_dir", "")
    if not download_dir:
        problems.append("Choose a folder for downloaded music.")
    else:
        try:
            Path(download_dir).expanduser().mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            problems.append(f"That download folder cannot be used: {exc}")
    if submitted.get("azuracast_enabled"):
        if not submitted.get("azuracast_api_base"):
            problems.append("Enter the AzuraCast base API address, for example http://127.0.0.1:12000/api.")
        if not submitted.get("azuracast_api_key"):
            problems.append("Enter an AzuraCast API key, or turn AzuraCast off to use local downloads only.")
    host = submitted.get("bind_host", "")
    no_password = clear_password or (not password and not PASSWORD_HASH)
    if require_passwords and not password:
        problems.append("Set a password.")
    elif host not in config.LOOPBACK_HOSTS and no_password:
        problems.append(
            "Set a password before listening on a network address, or bind to 127.0.0.1 "
            "so only this computer can reach the program."
        )
    if require_passwords and not admin_password:
        problems.append("Set an admin password.")
    return problems


@app.post("/setup/tools/update")
def update_tools():
    """Check for and install newer yt-dlp and musicdl releases now."""
    require_csrf()
    try:
        results = tools.update_all(SETTINGS)
    except Exception as exc:
        flash(f"The tools could not be updated: {exc}")
        return redirect(url_for("preferences"))
    tools.record_update_check(STATE_DIR, results)
    save_tool_paths(results)
    for item in results:
        flash(str(item.get("message") or ""))
    return redirect(url_for("preferences"))


@app.post("/setup/tools")
def setup_tools():
    """Fetch yt-dlp and ffmpeg into the program folder."""
    if not config.is_configured(SETTINGS):
        pass
    else:
        require_csrf()
    if not tools.can_auto_install():
        flash("Install yt-dlp and ffmpeg with your system package manager, then reload this page.")
        return redirect(url_for("setup" if not config.is_configured(SETTINGS) else "preferences"))
    try:
        updates = tools.install_missing(SETTINGS)
    except Exception as exc:
        flash(f"The download tools could not be installed: {exc}")
        return redirect(url_for("setup" if not config.is_configured(SETTINGS) else "preferences"))
    if updates:
        remember_settings({**SETTINGS, **updates})
        flash("The download tools are installed.")
    else:
        flash("The download tools were already available.")
    return redirect(url_for("setup" if not config.is_configured(SETTINGS) else "preferences"))


# Keep direct playlist and channel URL support active for every entry point,
# including run.py, Waitress, Gunicorn, and Flask's application loader.
from .youtube_collections import install as _install_youtube_collection_support

_install_youtube_collection_support()


apply_settings()

worker = None
updater = None
if not WORKER_DISABLED:
    worker = threading.Thread(target=worker_loop, name="download-worker", daemon=True)
    worker.start()
    # Keeping yt-dlp and musicdl current is what stops downloads failing against sites
    # that have moved on. It runs beside the worker so a check never delays a download.
    updater = threading.Thread(target=tool_update_loop, name="tool-updater", daemon=True)
    updater.start()
